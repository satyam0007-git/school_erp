import calendar
import io
import json
import re
import urllib.error
import urllib.request
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Max, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .forms import FeePaymentForm, StudentForm, TeacherForm, SalaryPaymentForm, SuperUserSettingsForm
from .models import (
    ExamFee, FeeCategory, FeePayment, FeeStructure, School, SchoolBillingPayment,
    SchoolClass, SchoolProfile, SchoolSessionRecord, Student, Teacher, SalaryPayment,
    SuperUserSettings, User, WhatsAppConfig, MONTH_CHOICES,
)
from .session_utils import (
    get_academic_session_choices, get_current_academic_session, get_session_start_year,
    get_session_months, MONTH_TO_CAL, CAL_TO_MONTH, format_academic_session,
)


def _get_school_billing_months(school):
    """Return (token, label) tuples for every month of the school's current academic session.

    Tokens are year-qualified ('april_2026') so they are unique across sessions.
    E.g. session '2026-27', start='april', end='march' →
         [('april_2026','April 2026'), …, ('march_2027','March 2027')]
    """
    profile = SchoolProfile.get_for_school(school)
    session = profile.current_academic_session          # e.g. '2026-27'
    try:
        session_start_year = int(session.split('-')[0])
    except (ValueError, IndexError):
        session_start_year = get_session_start_year(timezone.localdate(), profile.billing_start_month)

    start_key = profile.billing_start_month             # e.g. 'april'
    end_key   = profile.billing_end_month               # e.g. 'march'
    start_cal = MONTH_TO_CAL[start_key]

    months = get_session_months(start_key, end_key)
    result = []
    for m in months:
        cal = MONTH_TO_CAL[m]
        token_year = session_start_year if cal >= start_cal else session_start_year + 1
        token = f"{m}_{token_year}"
        label = f"{calendar.month_name[cal]} {token_year}"
        result.append((token, label))
    return result




# ── Decorators ────────────────────────────────────────────────────────────────

def super_only(fn):
    """Allow only superusers; always operates on the main (no-subdomain) domain."""
    @wraps(fn)
    def wrap(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_super():
            return redirect('login')
        return fn(request, *args, **kwargs)
    return wrap


def school_only(fn):
    """
    Allow only authenticated school admins.

    Subdomain enforcement:
    - If the request arrives on a school subdomain, the logged-in user's school
      must match that tenant.  If it doesn't, redirect the user to their own
      school's subdomain (or the generic login if their school has no subdomain).
    - Superusers are always bounced to the super dashboard.
    """
    @wraps(fn)
    def wrap(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')

        if request.user.is_super():
            return redirect('super_dashboard')

        tenant = getattr(request, 'tenant', None)
        if tenant is not None and request.user.school_id != tenant.pk:
            # User is logged in but belongs to a different school's subdomain.
            user_school = request.user.school
            if user_school and user_school.subdomain:
                return redirect(user_school.get_tenant_url() + '/school/')
            return redirect('login')

        return fn(request, *args, **kwargs)
    return wrap


# ── Auth ──────────────────────────────────────────────────────────────────────

def login_view(request):
    """
    Single login endpoint for all roles.

    Main domain  → superuser login form (login.html)
    Subdomain    → school admin login form (school/login.html), tenant-locked
    """
    tenant = getattr(request, 'tenant', None)

    if request.user.is_authenticated:
        if request.user.is_super():
            return redirect('super_dashboard')
        if tenant and request.user.school_id != tenant.pk:
            user_school = request.user.school
            if user_school and user_school.subdomain:
                return redirect(user_school.get_tenant_url() + '/school/')
        return redirect('school_dashboard')

    if request.method == 'POST':
        user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
        if user:
            if tenant:
                # School subdomain — reject superusers and wrong-school credentials
                if user.is_super():
                    messages.error(request, 'Super admin must log in from the main portal.')
                elif user.school_id != tenant.pk:
                    messages.error(request, 'These credentials do not belong to this school portal.')
                else:
                    login(request, user)
                    return redirect('school_dashboard')
            else:
                # Main domain — accept anyone, redirect to correct dashboard
                login(request, user)
                return redirect('dashboard')
        else:
            messages.error(request, 'Invalid credentials.')

    app_settings = SuperUserSettings.get_solo()
    template = 'school/login.html' if tenant else 'login.html'
    return render(request, template, {'settings': app_settings, 'tenant': tenant})


@login_required
def dashboard(request):
    if request.user.is_super():
        return redirect('super_dashboard')
    else:
        return redirect('school_dashboard')


def logout_view(request):
    logout(request)
    return redirect('login')


def _username_from_name(name):
    base = re.sub(r'[^\w]', '_', name.strip().lower())
    base = re.sub(r'_+', '_', base).strip('_')
    username = base
    counter = 2
    while User.objects.filter(username=username).exists():
        username = f'{base}_{counter}'
        counter += 1
    return username


@super_only
def school_add(request):
    session_choices = get_academic_session_choices(past_years=2, future_years=10)
    if request.method == 'POST':
        name = request.POST['name'].strip()
        password = request.POST['password']
        username = request.POST.get('username', '').strip().lower()
        if not name or not password or not username:
            messages.error(request, 'School name, username and password are required.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" is already taken. Please choose another.')
        else:
            valid_months = {m for m, _ in MONTH_CHOICES}
            billing_start = request.POST.get('billing_start_month', 'april')
            billing_end   = request.POST.get('billing_end_month', 'march')
            session       = request.POST.get('current_academic_session', session_choices[0][0])
            if billing_start not in valid_months:
                billing_start = 'april'
            if billing_end not in valid_months:
                billing_end = 'march'
            subdomain = request.POST.get('subdomain', '').strip().lower()
            if subdomain and School.objects.filter(subdomain=subdomain).exists():
                messages.error(request, f'Subdomain "{subdomain}" is already taken.')
                return render(request, 'superuser/school_add.html', {
                    'session_choices': session_choices,
                    'month_choices': MONTH_CHOICES,
                })
            school = School.objects.create(
                name=name,
                subdomain=subdomain or None,
                phone=request.POST.get('phone', ''),
                email=request.POST.get('email', ''),
                address=request.POST.get('address', ''),
                fee_per_student=request.POST.get('fee_per_student') or 0,
                logo=request.FILES.get('logo'),
            )
            SchoolProfile.objects.create(
                school=school,
                current_academic_session=session,
                billing_start_month=billing_start,
                billing_end_month=billing_end,
            )
            SchoolSessionRecord.objects.create(
                school=school,
                academic_session=session,
                session_start_month=SchoolProfile.objects.get(school=school).session_start_month,
                session_end_month=SchoolProfile.objects.get(school=school).session_end_month,
                billing_start_month=billing_start,
                billing_end_month=billing_end,
            )
            User.objects.create_user(username=username, password=password, role='school_admin', school=school)
            portal_url = school.get_tenant_url() if school.subdomain else '(no subdomain set)'
            messages.success(
                request,
                f'School "{name}" created. Admin username: {username}. Portal: {portal_url}'
            )
            return redirect('super_dashboard')
    from django.conf import settings as _s
    return render(request, 'superuser/school_add.html', {
        'session_choices': session_choices,
        'month_choices': MONTH_CHOICES,
        'base_domain': getattr(_s, 'TENANT_BASE_DOMAIN', 'localhost'),
    })


@super_only
def school_edit(request, pk):
    from django.conf import settings as _s
    base_domain = getattr(_s, 'TENANT_BASE_DOMAIN', 'localhost')
    school = get_object_or_404(School, pk=pk)
    admin_user = User.objects.filter(school=school, role='school_admin').first()
    profile = SchoolProfile.get_for_school(school)
    session_choices = get_academic_session_choices(past_years=2, future_years=10)
    valid_months = {m for m, _ in MONTH_CHOICES}

    def _render(extra=None):
        ctx = {
            'school': school, 'admin_user': admin_user, 'base_domain': base_domain,
            'profile': profile, 'session_choices': session_choices, 'month_choices': MONTH_CHOICES,
        }
        if extra:
            ctx.update(extra)
        return render(request, 'superuser/school_form.html', ctx)

    if request.method == 'POST':
        school.name = request.POST['name']
        school.phone = request.POST.get('phone', '')
        school.email = request.POST.get('email', '')
        school.address = request.POST.get('address', '')
        school.is_active = 'is_active' in request.POST
        school.fee_per_student = request.POST.get('fee_per_student') or 0
        new_logo = request.FILES.get('logo')
        if new_logo:
            school.logo = new_logo
        elif request.POST.get('clear_logo') == '1':
            school.logo = None
        new_subdomain = request.POST.get('subdomain', '').strip().lower() or None
        if new_subdomain != school.subdomain:
            if new_subdomain and School.objects.filter(subdomain=new_subdomain).exclude(pk=school.pk).exists():
                messages.error(request, f'Subdomain "{new_subdomain}" is already taken by another school.')
                return _render()
            school.subdomain = new_subdomain
        school.save()

        # Academic session
        old_session = profile.current_academic_session
        new_session = request.POST.get('current_academic_session', old_session)
        billing_start = request.POST.get('billing_start_month', profile.billing_start_month)
        billing_end = request.POST.get('billing_end_month', profile.billing_end_month)
        if billing_start not in valid_months:
            billing_start = 'april'
        if billing_end not in valid_months:
            billing_end = 'march'
        profile.current_academic_session = new_session
        profile.billing_start_month = billing_start
        profile.billing_end_month = billing_end
        profile.save()

        # Treat session change as a correction: delete old record, upsert new one.
        if old_session != new_session:
            SchoolSessionRecord.objects.filter(
                school=school, academic_session=old_session
            ).delete()
        SchoolSessionRecord.objects.update_or_create(
            school=school, academic_session=new_session,
            defaults={
                'session_start_month': profile.session_start_month,
                'session_end_month': profile.session_end_month,
                'billing_start_month': billing_start,
                'billing_end_month': billing_end,
            },
        )

        if admin_user:
            new_username = request.POST.get('admin_username', '').strip()
            new_password = request.POST.get('admin_password', '').strip()
            if new_username and new_username != admin_user.username:
                if User.objects.filter(username=new_username).exclude(pk=admin_user.pk).exists():
                    messages.error(request, 'Admin username already taken.')
                    return _render()
                admin_user.username = new_username
            if new_password:
                admin_user.set_password(new_password)
            admin_user.save()
        messages.success(request, 'School updated.')
        return redirect('super_dashboard')
    return _render()


@super_only
def school_delete(request, pk):
    school = get_object_or_404(School, pk=pk)
    if request.method == 'POST':
        FeePayment.objects.filter(school=school).delete()
        FeeStructure.objects.filter(school_class__school=school).delete()
        Student.objects.filter(school=school).delete()
        school.delete()
        messages.success(request, 'School deleted.')
        return redirect('super_dashboard')
    return render(request, 'confirm_delete.html', {'name': school.name})


@super_only
def user_add(request):
    if request.method == 'POST':
        username = request.POST['username'].strip()
        password = request.POST['password']
        school_id = request.POST.get('school')
        if not username or not password or not school_id:
            messages.error(request, 'All fields required.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Username taken.')
        else:
            school = get_object_or_404(School, pk=school_id)
            User.objects.create_user(username=username, password=password, role='school_admin', school=school)
            messages.success(request, 'User created.')
            return redirect('super_dashboard')
    return redirect('super_dashboard')


@super_only
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'Deleted.')
        return redirect('super_dashboard')
    return render(request, 'confirm_delete.html', {'name': user.username})


# ── School Dashboard ──────────────────────────────────────────────────────────

@school_only
def school_dashboard(request):
    return redirect('student_list')


# ── Admission ─────────────────────────────────────────────────────────────────

@school_only
def student_list(request):
    school = request.user.school
    qs = Student.objects.filter(school=school).select_related('school_class')
    q = request.GET.get('q', '').strip()
    class_id = request.GET.get('class')
    status = request.GET.get('status')
    profile = SchoolProfile.get_for_school(school)
    # Default to current session unless the user explicitly submitted a session filter
    if 'session' in request.GET:
        session = request.GET.get('session', '').strip()
    else:
        session = profile.current_academic_session
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(roll_number__icontains=q) | Q(father_name__icontains=q))
    if class_id:
        qs = qs.filter(school_class_id=class_id)
    if status:
        qs = qs.filter(status=status)
    if session:
        qs = qs.filter(academic_session=session)

    # KPIs derived from the filtered queryset (not all students)
    filtered_total = qs.count()
    filtered_active = qs.filter(status='active').count()
    filtered_promoted = qs.filter(status='promoted').count()
    filtered_inactive = qs.filter(status='inactive').count()
    filtered_fail = qs.filter(status='fail').count()

    filtered_class_wise = (
        qs.values('school_class__name')
        .annotate(count=Count('id'))
        .order_by('school_class__name')
    )

    all_students = Student.objects.filter(school=school).select_related('school_class')
    total_students = all_students.count()

    class_wise = (
        all_students.values('school_class__name')
        .annotate(count=Count('id'))
        .order_by('school_class__name')
    )
    session_wise = (
        all_students.values('academic_session')
        .annotate(count=Count('id'))
        .order_by('-academic_session')
    )

    return render(request, 'school/admission/student_list.html', {
        'students': qs[:50],
        'classes': SchoolClass.objects.filter(school=school),
        'session_choices': get_academic_session_choices(past_years=2, future_years=10),
        'status_choices': Student.STATUS_CHOICES,
        'current_session': profile.current_academic_session,
        'selected_session': session,
        'session_start_month': profile.session_start_month.capitalize(),
        'session_end_month': profile.session_end_month.capitalize(),
        'total_students': total_students,
        'class_wise': class_wise,
        'session_wise': session_wise,
        'filtered_total': filtered_total,
        'filtered_active': filtered_active,
        'filtered_promoted': filtered_promoted,
        'filtered_inactive': filtered_inactive,
        'filtered_fail': filtered_fail,
        'filtered_class_wise': filtered_class_wise,
        'selected_class': class_id or '',
        'selected_status': status or '',
        'selected_q': q,
        'has_filters': bool(q or class_id or status or (session and session != profile.current_academic_session)),
    })


@school_only
def student_create(request):
    school = request.user.school
    form = StudentForm(request.POST or None, school=school)
    if request.method == 'POST' and form.is_valid():
        student = form.save(commit=False)
        student.school = school
        student.save()
        discount_months = form.cleaned_data.get('discount_months')
        if discount_months and discount_months > 0:
            messages.success(request, f'Student admitted. First {discount_months} month(s) discounted.')
        else:
            messages.success(request, 'Student admitted.')
        return redirect('student_list')
    return render(request, 'school/admission/student_form.html', {'form': form})


@school_only
def student_edit(request, pk):
    school = request.user.school
    student = get_object_or_404(Student, pk=pk, school=school)
    form = StudentForm(request.POST or None, instance=student, school=school)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Student updated.')
        return redirect('student_list')
    return render(request, 'school/admission/student_form.html', {'form': form, 'object': student})


@school_only
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk, school=request.user.school)
    if request.method == 'POST':
        student.delete()
        messages.success(request, 'Student deleted.')
        return redirect('student_list')
    return render(request, 'school/admission/confirm_delete.html', {'object': student})


@school_only
def student_promote(request, pk):
    if request.method != 'POST':
        return redirect('student_list')
    school = request.user.school
    student = get_object_or_404(Student, pk=pk, school=school, status=Student.STATUS_ACTIVE)

    classes = list(SchoolClass.objects.filter(school=school).order_by('name'))
    current_idx = next((i for i, c in enumerate(classes) if c.pk == student.school_class_id), None)

    if current_idx is None or current_idx >= len(classes) - 1:
        messages.error(request, f'{student.name} is already in the last class and cannot be promoted further.')
        return redirect('student_list')

    next_class = classes[current_idx + 1]

    try:
        start_year = int(student.academic_session.split('-')[0])
        next_session = f"{start_year + 1}-{(start_year + 2) % 100:02d}"
    except (ValueError, IndexError, AttributeError):
        next_session = get_current_academic_session(timezone.localdate())

    old_class = student.school_class.name
    old_session = student.academic_session

    with transaction.atomic():
        student.status = Student.STATUS_PROMOTED
        student.save(update_fields=['status', 'updated_at'])

        promoted_student = Student.objects.create(
            school=student.school,
            school_class=next_class,
            name=student.name,
            date_of_birth=student.date_of_birth,
            academic_session=next_session,
            status=Student.STATUS_ACTIVE,
            father_name=student.father_name,
            father_phone=student.father_phone,
            address=student.address,
            admission_date=student.admission_date,
            transport_opted=student.transport_opted,
            transport_amount=student.transport_amount,
        )

    messages.success(
        request,
        f'{student.name} promoted from {old_class} ({old_session}) → {promoted_student.school_class.name} ({promoted_student.academic_session}).',
    )
    return redirect('student_list')


@school_only
def student_fail(request, pk):
    if request.method != 'POST':
        return redirect('student_list')
    school = request.user.school
    student = get_object_or_404(Student, pk=pk, school=school, status=Student.STATUS_ACTIVE)

    try:
        start_year = int(student.academic_session.split('-')[0])
        next_session = f"{start_year + 1}-{(start_year + 2) % 100:02d}"
    except (ValueError, IndexError, AttributeError):
        next_session = get_current_academic_session(timezone.localdate())

    old_class = student.school_class.name
    old_session = student.academic_session

    with transaction.atomic():
        student.status = Student.STATUS_FAIL
        student.save(update_fields=['status', 'updated_at'])

        failed_student = Student.objects.create(
            school=student.school,
            school_class=student.school_class,  # Same class, not next class
            name=student.name,
            date_of_birth=student.date_of_birth,
            academic_session=next_session,
            status=Student.STATUS_ACTIVE,
            father_name=student.father_name,
            father_phone=student.father_phone,
            address=student.address,
            admission_date=student.admission_date,
            transport_opted=student.transport_opted,
            transport_amount=student.transport_amount,
        )

    messages.success(
        request,
        f'{student.name} marked as failed and retained in {old_class} ({next_session}).',
    )
    return redirect('student_list')


@school_only
def student_transfer(request, pk):
    if request.method != 'POST':
        return redirect('student_list')
    student = get_object_or_404(Student, pk=pk, school=request.user.school, status=Student.STATUS_ACTIVE)
    student.status = Student.STATUS_INACTIVE
    student.save(update_fields=['status', 'updated_at'])
    messages.success(request, f'{student.name} marked as transferred. Record kept for {student.academic_session}.')
    return redirect('student_list')


# ── Fee Submission ────────────────────────────────────────────────────────────

def _get_monthly_fee_total(student, school, session=None):
    if session is None:
        profile = SchoolProfile.get_for_school(school)
        session = profile.current_academic_session
    return FeeStructure.objects.filter(
        school_class=student.school_class,
        fee_category__school=school,
        fee_category__is_active=True,
        academic_session=session,
        frequency=FeeStructure.FREQUENCY_MONTHLY,
    ).aggregate(t=Sum('amount'))['t'] or Decimal('0.00')


def _get_monthly_transport_fee(student):
    if student.transport_opted and student.transport_amount:
        return student.transport_amount
    return Decimal('0.00')


def _get_available_advance(student, session):
    """Net advance balance available for a student in a given session."""
    result = FeePayment.objects.filter(
        student=student, academic_session=session
    ).aggregate(total_advance=Sum('advance_balance'), total_used=Sum('advance_used'))
    return (result['total_advance'] or Decimal('0.00')) - (result['total_used'] or Decimal('0.00'))


def _get_discount_covered_tokens(student, school, profile):
    """Return month tokens removed from payable list due to admission discount months."""
    n = student.discount_months
    if not n or n <= 0:
        return set()
    session_months = get_session_months(profile.session_start_month, profile.session_end_month)
    has_transport = bool(student.transport_opted and student.transport_amount)
    tokens = set()
    for month in session_months[:n]:
        tokens.add(month)
        if has_transport:
            tokens.add(f'{month}_transport')
    return tokens


def _distribute_lump_sum(student, school, cash_amount, advance_available, profile):
    """Distribute a lump-sum amount across unpaid months/fees automatically.

    Priority order per month: tuition fee → transport fee.
    After all months exhausted: exam fees (each tried independently).
    Any amount that cannot cover a full fee item becomes advance balance.
    """
    session = profile.current_academic_session
    session_months = get_session_months(profile.session_start_month, profile.session_end_month)

    cash_amount = Decimal(str(cash_amount))
    advance_available = Decimal(str(advance_available))
    remaining = cash_amount + advance_available

    monthly_fee = _get_monthly_fee_total(student, school, session)
    transport_fee = _get_monthly_transport_fee(student)
    has_transport = bool(student.transport_opted and student.transport_amount)

    # Collect already-paid tokens (FeePayments + admission discount)
    paid_tokens = set()
    paid_exam_names = set()
    session_qs = FeePayment.objects.filter(student=student, academic_session=session)
    for month_list in session_qs.values_list('payment_months', flat=True):
        if isinstance(month_list, list):
            paid_tokens.update(str(t) for t in month_list)
    for ep in session_qs.values_list('exam_fee_items', flat=True):
        for item in (ep or []):
            name = item.get('name') or item.get('exam_name') or str(item)
            paid_exam_names.add(name)
    paid_tokens |= _get_discount_covered_tokens(student, school, profile)

    paid_month_tokens = []
    exam_fee_items = []
    transport_total = Decimal('0.00')

    # Distribute month by month: tuition first, then transport
    if monthly_fee > 0:
        for month in session_months:
            if month in paid_tokens:
                continue
            if remaining < monthly_fee:
                break  # insufficient for full tuition — remaining becomes advance
            paid_month_tokens.append(month)
            remaining -= monthly_fee

            transport_token = f'{month}_transport'
            if has_transport and transport_token not in paid_tokens:
                if remaining < transport_fee:
                    break  # insufficient for transport — remaining becomes advance
                paid_month_tokens.append(transport_token)
                remaining -= transport_fee
                transport_total += transport_fee

    # Distribute to exam fees (each attempted independently)
    for ef in ExamFee.objects.filter(
        school=school, school_class=student.school_class, academic_session=session
    ).order_by('exam_name'):
        if ef.exam_name in paid_exam_names:
            continue
        if remaining >= ef.amount:
            exam_fee_items.append({'name': ef.exam_name, 'amount': str(ef.amount)})
            remaining -= ef.amount

    advance_balance = remaining
    gross_amount = (cash_amount + advance_available) - advance_balance

    return {
        'paid_month_tokens': paid_month_tokens,
        'exam_fee_items': exam_fee_items,
        'transport_total': transport_total,
        'gross_amount': gross_amount,
        'advance_balance': advance_balance,
    }


def _get_unpaid_month_options(student, school):
    profile = SchoolProfile.get_for_school(school)
    target_session = profile.current_academic_session
    session_months = get_session_months(profile.session_start_month, profile.session_end_month)

    has_monthly = FeeStructure.objects.filter(
        school_class=student.school_class,
        fee_category__school=school,
        fee_category__is_active=True,
        academic_session=target_session,
        frequency=FeeStructure.FREQUENCY_MONTHLY,
    ).exists()
    has_transport = bool(student.transport_opted and student.transport_amount)

    valid_months = set(session_months)
    valid_tokens = valid_months | {f'{v}_transport' for v in valid_months}

    paid_tokens = set()
    session_payments_qs = FeePayment.objects.filter(
        student=student, academic_session=target_session
    )
    for month_list in session_payments_qs.values_list('payment_months', flat=True):
        if isinstance(month_list, list):
            for t in month_list:
                if str(t) in valid_tokens:
                    paid_tokens.add(str(t))

    # Also exclude months already covered by admission discount
    paid_tokens |= _get_discount_covered_tokens(student, school, profile)

    month_label_map = dict(MONTH_CHOICES)
    options = []
    for value in session_months:
        label = month_label_map[value]
        if has_monthly and value not in paid_tokens:
            options.append({'value': value, 'label': label, 'group': 'Monthly Fee'})
        transport_token = f'{value}_transport'
        if has_transport and transport_token not in paid_tokens:
            options.append({'value': transport_token, 'label': f'{label} (Transport)', 'group': 'Transport'})

    # Add unpaid exam fees for student's class
    paid_exam_names = set()
    for ep in session_payments_qs.values_list('exam_fee_items', flat=True):
        for item in (ep or []):
            name = item.get('name') or item.get('exam_name') or str(item)
            paid_exam_names.add(name)

    for ef in ExamFee.objects.filter(school=school, school_class=student.school_class, academic_session=target_session):
        if ef.exam_name not in paid_exam_names:
            options.append({'value': f'exam_{ef.pk}', 'label': f'{ef.exam_name} — ₹{ef.amount:,.0f}', 'group': 'Exam Fee'})

    return options


def _get_edit_month_options(payment, school):
    """Month options when editing an existing FeePayment.

    Returns months already in this payment (pre-selected) plus months not
    yet paid by any *other* payment for the same student/session.
    """
    student = payment.student
    profile = SchoolProfile.get_for_school(school)
    target_session = profile.current_academic_session
    session_months = get_session_months(profile.session_start_month, profile.session_end_month)

    has_monthly = FeeStructure.objects.filter(
        school_class=student.school_class,
        fee_category__school=school,
        fee_category__is_active=True,
        academic_session=target_session,
        frequency=FeeStructure.FREQUENCY_MONTHLY,
    ).exists()
    has_transport = bool(student.transport_opted and student.transport_amount)

    valid_months = set(session_months)
    valid_tokens = valid_months | {f'{v}_transport' for v in valid_months}

    # Tokens paid by OTHER payments (exclude the payment being edited)
    other_paid_tokens = set()
    for month_list in FeePayment.objects.filter(
        student=student, academic_session=target_session
    ).exclude(pk=payment.pk).values_list('payment_months', flat=True):
        if isinstance(month_list, list):
            for t in month_list:
                if str(t) in valid_tokens:
                    other_paid_tokens.add(str(t))

    this_tokens = set(str(t) for t in (payment.payment_months or []))

    month_label_map = dict(MONTH_CHOICES)
    options = []
    for value in session_months:
        label = month_label_map[value]
        if has_monthly and (value in this_tokens or value not in other_paid_tokens):
            options.append({
                'value': value,
                'label': label,
                'selected': value in this_tokens,
                'group': 'Monthly Fee',
            })
        transport_token = f'{value}_transport'
        if has_transport and (transport_token in this_tokens or transport_token not in other_paid_tokens):
            options.append({
                'value': transport_token,
                'label': f'{label} (Transport)',
                'selected': transport_token in this_tokens,
                'group': 'Transport',
            })

    # Exam fees: available if not paid in another payment, or already in this payment
    paid_exam_names_other = set()
    for ep in FeePayment.objects.filter(
        student=student, academic_session=target_session
    ).exclude(pk=payment.pk).values_list('exam_fee_items', flat=True):
        for item in (ep or []):
            name = item.get('name') or item.get('exam_name') or str(item)
            paid_exam_names_other.add(name)

    this_exam_names = set()
    for item in (payment.exam_fee_items or []):
        name = item.get('name') or item.get('exam_name') or str(item)
        this_exam_names.add(name)

    for ef in ExamFee.objects.filter(school=school, school_class=student.school_class, academic_session=target_session):
        if ef.exam_name not in paid_exam_names_other:
            options.append({
                'value': f'exam_{ef.pk}',
                'label': f'{ef.exam_name} — ₹{ef.amount:,.0f}',
                'selected': ef.exam_name in this_exam_names,
                'group': 'Exam Fee',
            })

    return options


@school_only
def payment_dashboard(request):
    school = request.user.school
    reference_date = timezone.localdate()
    profile = SchoolProfile.get_for_school(school)
    profile_session = profile.current_academic_session

    q = request.GET.get('q', '').strip()
    class_id = request.GET.get('class')
    payment_status = request.GET.get('payment_status', '')
    selected_session = request.GET.get('session', '').strip() or profile_session

    session_choices = get_academic_session_choices(
        past_years=2, future_years=10,
        session_start_month=profile.session_start_month,
    )

    students = Student.objects.filter(school=school, status=Student.STATUS_ACTIVE).select_related('school_class')
    if q:
        students = students.filter(
            Q(name__icontains=q) | Q(father_name__icontains=q) | Q(roll_number__icontains=q)
        )
    if class_id:
        students = students.filter(school_class_id=class_id)

    session_months = get_session_months(profile.session_start_month, profile.session_end_month)
    current_month_key = CAL_TO_MONTH[reference_date.month]

    # Determine how many months are "due" for the selected session
    if selected_session == profile_session:
        if current_month_key in session_months:
            default_months_due = session_months.index(current_month_key) + 1
        else:
            default_months_due = len(session_months)
    elif selected_session < profile_session:
        default_months_due = len(session_months)  # past session — all months due
    else:
        default_months_due = 0  # future session — nothing due yet

    month_label_map = {v: l for v, l in MONTH_CHOICES}

    # Latest payment ID per student (for Edit button)
    latest_payment_by_student = dict(
        FeePayment.objects.filter(school=school, academic_session=selected_session)
        .values('student_id')
        .annotate(latest_id=Max('id'))
        .values_list('student_id', 'latest_id')
    )

    # Collect paid months per student for selected session in one query
    session_payments = FeePayment.objects.filter(
        school=school,
        academic_session=selected_session,
    ).values('student_id', 'payment_months', 'exam_fee_items', 'payment_date')
    paid_months_by_student = {}
    exam_items_by_student = {}
    for p in session_payments:
        sid = p['student_id']
        paid_months_by_student.setdefault(sid, [])
        exam_items_by_student.setdefault(sid, [])
        date_str = p['payment_date'].strftime('%d %b %Y') if p['payment_date'] else '—'
        for m in (p['payment_months'] or []):
            m = str(m)
            if m.endswith('_transport'):
                base = m.replace('_transport', '')
                label = month_label_map.get(base, base)
                paid_months_by_student[sid].append({'label': label, 'transport': True, 'date': date_str})
            else:
                label = month_label_map.get(m, m)
                paid_months_by_student[sid].append({'label': label, 'transport': False, 'date': date_str})
        for item in (p['exam_fee_items'] or []):
            name = item.get('name') or item.get('exam_name') or str(item)
            entry = {'name': name, 'date': date_str}
            if not any(e['name'] == name for e in exam_items_by_student[sid]):
                exam_items_by_student[sid].append(entry)

    payment_data = []
    for student in students.order_by('school_class__name', 'name'):
        months_due = default_months_due

        monthly_fee = _get_monthly_fee_total(student, school, session=selected_session)
        transport_fee = _get_monthly_transport_fee(student)
        total_needed = (monthly_fee + transport_fee) * months_due
        total_paid = FeePayment.objects.filter(
            student=student,
            academic_session=selected_session,
        ).aggregate(t=Sum('amount_paid'))['t'] or Decimal('0.00')
        balance = max(total_needed - total_paid, Decimal('0.00'))

        paid_months = paid_months_by_student.get(student.id, [])

        payment_data.append({
            'student_id': student.id,
            'student_name': student.name,
            'roll_number': student.roll_number,
            'father_name': student.father_name,
            'student_class': str(student.school_class),
            'father_phone': student.father_phone,
            'total_amount_paid': total_paid,
            'total_payment_need_to_pay_till_month': total_needed,
            'balance_payment': balance,
            'is_paid_up': balance <= Decimal('0.00'),
            'paid_months': paid_months,
            'exam_items': exam_items_by_student.get(student.id, []),
            'latest_payment_id': latest_payment_by_student.get(student.id),
        })

    if payment_status == 'pending':
        payment_data = [p for p in payment_data if p['balance_payment'] > 0]
    elif payment_status == 'paid':
        payment_data = [p for p in payment_data if p['balance_payment'] <= 0]

    total_due = sum(p['total_payment_need_to_pay_till_month'] for p in payment_data)
    total_collected = sum(p['total_amount_paid'] for p in payment_data)
    total_pending = max(total_due - total_collected, Decimal('0.00'))

    paginator = Paginator(payment_data, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Fee data range label (e.g. "April 2026 → May 2026")
    session_start_year = int(selected_session[:4]) if selected_session and len(selected_session) >= 4 else reference_date.year
    session_start_cal = MONTH_TO_CAL[profile.session_start_month]
    _range_start_key = session_months[0] if session_months else profile.session_start_month
    _range_end_key   = session_months[default_months_due - 1] if default_months_due > 0 else session_months[0]
    _sy = lambda key: session_start_year if MONTH_TO_CAL[key] >= session_start_cal else session_start_year + 1
    fee_range_label = f"{_range_start_key.capitalize()} {_sy(_range_start_key)} → {_range_end_key.capitalize()} {_sy(_range_end_key)}"

    wa_config, _ = WhatsAppConfig.objects.get_or_create(school=school)
    has_filters = bool(
        q or class_id or payment_status
        or (selected_session and selected_session != profile_session)
    )
    return render(request, 'school/fees/payment_dashboard.html', {
        'payment_summary': page_obj,
        'page_obj': page_obj,
        'is_paginated': paginator.num_pages > 1,
        'classes': SchoolClass.objects.filter(school=school),
        'session_choices': session_choices,
        'selected_session': selected_session,
        'selected_class': class_id or '',
        'selected_q': q,
        'selected_payment_status': payment_status,
        'has_filters': has_filters,
        'total_students': len(payment_data),
        'total_due': total_due,
        'total_collected': total_collected,
        'total_pending': total_pending,
        'wa_enabled': wa_config.is_active,
        'fee_range_label': fee_range_label,
    })


def _build_fee_dashboard_data(request):
    """Shared data-building logic for export/print of the fee submission dashboard."""
    school = request.user.school
    reference_date = timezone.localdate()
    profile = SchoolProfile.get_for_school(school)
    profile_session = profile.current_academic_session

    q = request.GET.get('q', '').strip()
    class_id = request.GET.get('class')
    payment_status = request.GET.get('payment_status', '')
    selected_session = request.GET.get('session', '').strip() or profile_session

    students = Student.objects.filter(school=school, status=Student.STATUS_ACTIVE).select_related('school_class')
    if q:
        students = students.filter(
            Q(name__icontains=q) | Q(father_name__icontains=q) | Q(roll_number__icontains=q)
        )
    if class_id:
        students = students.filter(school_class_id=class_id)

    session_months = get_session_months(profile.session_start_month, profile.session_end_month)
    current_month_key = CAL_TO_MONTH[reference_date.month]

    if selected_session == profile_session:
        if current_month_key in session_months:
            default_months_due = session_months.index(current_month_key) + 1
        else:
            default_months_due = len(session_months)
    elif selected_session < profile_session:
        default_months_due = len(session_months)
    else:
        default_months_due = 0

    payment_data = []
    for student in students.order_by('school_class__name', 'name'):
        monthly_fee = _get_monthly_fee_total(student, school, session=selected_session)
        transport_fee = _get_monthly_transport_fee(student)
        total_needed = (monthly_fee + transport_fee) * default_months_due
        total_paid = FeePayment.objects.filter(
            student=student,
            academic_session=selected_session,
        ).aggregate(t=Sum('amount_paid'))['t'] or Decimal('0.00')
        balance = max(total_needed - total_paid, Decimal('0.00'))
        payment_data.append({
            'student_name': student.name,
            'roll_number': student.roll_number,
            'father_name': student.father_name,
            'student_class': str(student.school_class),
            'total_amount_paid': total_paid,
            'total_payment_need_to_pay_till_month': total_needed,
            'balance_payment': balance,
            'is_paid_up': balance <= Decimal('0.00'),
        })

    if payment_status == 'pending':
        payment_data = [p for p in payment_data if p['balance_payment'] > 0]
    elif payment_status == 'paid':
        payment_data = [p for p in payment_data if p['balance_payment'] <= 0]

    total_due = sum(p['total_payment_need_to_pay_till_month'] for p in payment_data)
    total_collected = sum(p['total_amount_paid'] for p in payment_data)
    total_pending = max(total_due - total_collected, Decimal('0.00'))

    session_start_year = int(selected_session[:4]) if selected_session and len(selected_session) >= 4 else reference_date.year
    session_start_cal = MONTH_TO_CAL[profile.session_start_month]
    _range_start_key = session_months[0] if session_months else profile.session_start_month
    _range_end_key   = session_months[default_months_due - 1] if default_months_due > 0 else session_months[0]
    _sy = lambda key: session_start_year if MONTH_TO_CAL[key] >= session_start_cal else session_start_year + 1
    fee_range_label = f"{_range_start_key.capitalize()} {_sy(_range_start_key)} → {_range_end_key.capitalize()} {_sy(_range_end_key)}"

    return {
        'payment_data': payment_data,
        'total_due': total_due,
        'total_collected': total_collected,
        'total_pending': total_pending,
        'selected_session': selected_session,
        'school': school,
        'profile': profile,
        'fee_range_label': fee_range_label,
    }


@school_only
def payment_dashboard_export(request):
    data = _build_fee_dashboard_data(request)
    payment_data = data['payment_data']
    school = data['school']
    selected_session = data['selected_session']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fee Submission'

    num_cols = 8  # number of table columns
    last_col = get_column_letter(num_cols)

    # Info rows — merged across all columns so long text doesn't widen column A
    ws.merge_cells(f'A1:{last_col}1')
    ws.merge_cells(f'A2:{last_col}2')
    ws.merge_cells(f'A3:{last_col}3')
    ws['A1'].value = f"School: {school.name}"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A2'].value = f"Session: {selected_session}"
    ws['A2'].font = Font(size=10, color='64748B')
    ws['A3'].value = f"Billing Period: {data['fee_range_label']}"
    ws['A3'].font = Font(size=10, color='0D9488')
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 6   # spacer
    ws.row_dimensions[5].height = 30

    headers = ['S.N', 'Student Name', "Father's Name", 'Class', 'Total Due (₹)', 'Paid (₹)', 'Balance (₹)', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        _style_header_cell(cell, '1E40AF')

    for idx, p in enumerate(payment_data, 6):
        row_data = [
            idx - 5,
            p['student_name'],
            p['father_name'],
            p['student_class'],
            float(p['total_payment_need_to_pay_till_month']),
            float(p['total_amount_paid']),
            float(p['balance_payment']),
            'Paid' if p['is_paid_up'] else 'Pending',
        ]
        ws.append(row_data)
        for cell in ws[idx]:
            _style_data_cell(cell, idx)

    # Summary row
    summary_row = ws.max_row + 1
    ws.cell(row=summary_row, column=1, value='TOTAL')
    ws.cell(row=summary_row, column=5, value=float(data['total_due']))
    ws.cell(row=summary_row, column=6, value=float(data['total_collected']))
    ws.cell(row=summary_row, column=7, value=float(data['total_pending']))
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=summary_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='EFF6FF')
        thin = Side(style='thin', color='BFDBFE')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    _auto_width(ws, skip_rows=4)
    ws.freeze_panes = 'A6'

    safe_name = re.sub(r'[^\w\-]', '_', school.name)
    filename = f"fee_submission_{safe_name}_{selected_session}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@school_only
def payment_dashboard_print(request):
    data = _build_fee_dashboard_data(request)
    return render(request, 'school/fees/print_fee_dashboard.html', {
        'payment_data': data['payment_data'],
        'total_due': data['total_due'],
        'total_collected': data['total_collected'],
        'total_pending': data['total_pending'],
        'selected_session': data['selected_session'],
        'school': data['school'],
        'total_students': len(data['payment_data']),
        'printed_on': timezone.localdate(),
        'fee_range_label': data['fee_range_label'],
    })


@school_only
def payment_create(request):
    school = request.user.school
    student_qs = Student.objects.filter(school=school, status=Student.STATUS_ACTIVE).select_related('school_class')

    q = request.GET.get('q', '').strip()
    class_id = request.GET.get('class')
    session = request.GET.get('session')
    filtered_qs = student_qs
    if q:
        filtered_qs = filtered_qs.filter(Q(name__icontains=q) | Q(roll_number__icontains=q) | Q(father_name__icontains=q))
    if class_id:
        filtered_qs = filtered_qs.filter(school_class_id=class_id)
    if session:
        filtered_qs = filtered_qs.filter(academic_session=session)

    has_active_filters = bool(q or class_id or session)
    paginator = Paginator(filtered_qs.order_by('name'), 10)
    filtered_page = paginator.get_page(request.GET.get('page'))

    if request.method == 'POST':
        payment_mode = request.POST.get('payment_mode', 'manual')
        form = FeePaymentForm(request.POST, student_queryset=student_qs, lump_sum_mode=(payment_mode == 'lump_sum'))

        if form.is_valid():
            with transaction.atomic():
                payment = form.save(commit=False)
                profile = SchoolProfile.get_for_school(school)
                payment.school = school
                payment.academic_session = profile.current_academic_session
                payment.collected_by = request.user

                if payment_mode == 'lump_sum':
                    try:
                        cash_amount = Decimal(str(request.POST.get('lump_sum_amount', '0')))
                    except InvalidOperation:
                        cash_amount = Decimal('0.00')

                    if cash_amount <= 0:
                        form.add_error(None, 'Enter a valid lump-sum amount greater than zero.')
                    else:
                        advance_available = _get_available_advance(payment.student, profile.current_academic_session)
                        result = _distribute_lump_sum(payment.student, school, cash_amount, advance_available, profile)

                        if not result['paid_month_tokens'] and not result['exam_fee_items']:
                            form.add_error(None, 'All fees are already paid for this student or no fee structure is set up.')
                        else:
                            payment.is_lump_sum = True
                            payment.payment_months = result['paid_month_tokens']
                            payment.exam_fee_items = result['exam_fee_items']
                            payment.transport_amount = result['transport_total']
                            payment.amount_paid = cash_amount
                            payment.gross_amount = result['gross_amount']
                            payment.advance_used = advance_available
                            payment.advance_balance = result['advance_balance']
                            payment.save()
                            messages.success(request, 'Lump-sum payment saved and auto-distributed.')
                            return redirect('payment_dashboard')
                else:
                    selected_items = form.cleaned_data['payment_months']
                    selected_months = [m for m in selected_items if not str(m).startswith('exam_')]
                    selected_exam_keys = [m for m in selected_items if str(m).startswith('exam_')]

                    unpaid_options = _get_unpaid_month_options(payment.student, school)
                    unpaid_values = {item['value'] for item in unpaid_options}
                    invalid = [m for m in selected_items if m not in unpaid_values]
                    if invalid:
                        form.add_error('payment_months', 'Some selected items are already paid.')
                    else:
                        transport_count = sum(1 for m in selected_months if str(m).endswith('_transport'))
                        base_count = len(selected_months) - transport_count
                        monthly_base = _get_monthly_fee_total(payment.student, school)
                        monthly_transport = _get_monthly_transport_fee(payment.student)

                        exam_fee_items = []
                        exam_total = Decimal('0.00')
                        if selected_exam_keys:
                            exam_pk_list = [int(k.replace('exam_', '')) for k in selected_exam_keys if k.replace('exam_', '').isdigit()]
                            for ef in ExamFee.objects.filter(pk__in=exam_pk_list, school=school):
                                exam_fee_items.append({'name': ef.exam_name, 'amount': str(ef.amount)})
                                exam_total += ef.amount

                        month_amount = (monthly_base * base_count) + (monthly_transport * transport_count)
                        payment.payment_months = selected_months
                        payment.exam_fee_items = exam_fee_items
                        payment.transport_amount = monthly_transport * transport_count
                        payment.amount_paid = month_amount + exam_total
                        payment.gross_amount = payment.amount_paid
                        payment.save()
                        messages.success(request, 'Payment saved.')
                        return redirect('payment_dashboard')
    else:
        student_id_param = request.GET.get('student_id')
        form = FeePaymentForm(student_queryset=student_qs, initial_student_id=student_id_param)

    params = request.GET.copy()
    params.pop('page', None)

    selected_student_id = request.GET.get('student_id') or (request.POST.get('student') if request.method == 'POST' else None)
    payment_history = []
    selected_student = None
    advance_available = Decimal('0.00')
    if selected_student_id:
        try:
            selected_student = student_qs.get(pk=selected_student_id)
            payment_history = FeePayment.objects.filter(student=selected_student).select_related('student', 'student__school_class').order_by('-payment_date', '-id')
            profile_for_advance = SchoolProfile.get_for_school(school)
            advance_available = _get_available_advance(selected_student, profile_for_advance.current_academic_session)
        except Student.DoesNotExist:
            pass

    # Compute unpaid months server-side so the initial render is always correct
    initial_month_options = []
    if selected_student:
        initial_month_options = _get_unpaid_month_options(selected_student, school)

    month_label_map = dict(MONTH_CHOICES)

    return render(request, 'school/fees/payment_form.html', {
        'form': form,
        'classes': SchoolClass.objects.filter(school=school),
        'session_choices': get_academic_session_choices(past_years=2, future_years=10),
        'filtered_students_page': filtered_page,
        'total_filtered_students': filtered_qs.count(),
        'has_active_filters': has_active_filters,
        'filter_querystring': params.urlencode(),
        'payment_history': payment_history,
        'selected_student': selected_student,
        'initial_month_options': initial_month_options,
        'month_label_map': month_label_map,
        'advance_available': advance_available,
    })


@school_only
def payment_detail(request, pk):
    payment = get_object_or_404(FeePayment, pk=pk, school=request.user.school)
    return render(request, 'school/fees/payment_detail.html', {'payment': payment})


@school_only
def payment_edit(request, pk):
    school = request.user.school
    payment = get_object_or_404(FeePayment, pk=pk, school=school)
    student = payment.student
    profile = SchoolProfile.get_for_school(school)
    errors = []

    if request.method == 'POST':
        payment_date_str = request.POST.get('payment_date', '').strip()
        selected_items = request.POST.getlist('payment_months')

        try:
            payment_date = date.fromisoformat(payment_date_str)
        except (ValueError, TypeError):
            errors.append('Invalid payment date.')
            payment_date = payment.payment_date

        if not selected_items:
            errors.append('Select at least one month or fee.')

        if not errors:
            session_months = get_session_months(profile.session_start_month, profile.session_end_month)
            valid_tokens = set(session_months) | {f'{v}_transport' for v in session_months}

            other_paid_tokens = set()
            for month_list in FeePayment.objects.filter(
                student=student, academic_session=payment.academic_session
            ).exclude(pk=pk).values_list('payment_months', flat=True):
                if isinstance(month_list, list):
                    for t in month_list:
                        if str(t) in valid_tokens:
                            other_paid_tokens.add(str(t))

            selected_months = [m for m in selected_items if not str(m).startswith('exam_')]
            selected_exam_keys = [m for m in selected_items if str(m).startswith('exam_')]

            invalid = [m for m in selected_months if m in other_paid_tokens]
            if invalid:
                errors.append('Some selected months are already paid in another payment.')
            else:
                with transaction.atomic():
                    transport_count = sum(1 for m in selected_months if str(m).endswith('_transport'))
                    base_count = len(selected_months) - transport_count
                    monthly_base = _get_monthly_fee_total(student, school)
                    monthly_transport = _get_monthly_transport_fee(student)

                    exam_fee_items = []
                    exam_total = Decimal('0.00')
                    if selected_exam_keys:
                        exam_pk_list = [int(k.replace('exam_', '')) for k in selected_exam_keys if k.replace('exam_', '').isdigit()]
                        for ef in ExamFee.objects.filter(pk__in=exam_pk_list, school=school):
                            exam_fee_items.append({'name': ef.exam_name, 'amount': str(ef.amount)})
                            exam_total += ef.amount

                    payment.payment_date = payment_date
                    payment.payment_months = selected_months
                    payment.exam_fee_items = exam_fee_items
                    payment.transport_amount = monthly_transport * transport_count
                    payment.amount_paid = (monthly_base * base_count) + (monthly_transport * transport_count) + exam_total
                    payment.gross_amount = payment.amount_paid
                    payment.save()
                    messages.success(request, 'Payment updated successfully.')
                    return redirect('payment_dashboard')

    month_options = _get_edit_month_options(payment, school)
    student_payments = FeePayment.objects.filter(
        student=student, school=school
    ).order_by('-payment_date', '-id')
    return render(request, 'school/fees/payment_edit.html', {
        'payment': payment,
        'student': student,
        'month_options': month_options,
        'errors': errors,
        'student_payments': student_payments,
    })


@school_only
def student_fee_structure_ajax(request):
    school = request.user.school
    student_id = request.GET.get('student_id')
    if not student_id:
        return JsonResponse({'items': []})
    student = get_object_or_404(Student, pk=student_id, school=school, status=Student.STATUS_ACTIVE)
    profile = SchoolProfile.get_for_school(school)
    structures = FeeStructure.objects.filter(
        school_class=student.school_class,
        fee_category__school=school,
        fee_category__is_active=True,
        academic_session=profile.current_academic_session,
    ).select_related('fee_category')
    items = [{'category_name': r.fee_category.name, 'amount': str(r.amount), 'frequency': r.frequency, 'is_transport': False} for r in structures]
    if student.transport_opted and student.transport_amount:
        items.append({'category_name': 'Transport Fee', 'amount': str(student.transport_amount), 'frequency': 'monthly', 'is_transport': True})
    exam_amounts = {
        f'exam_{ef.pk}': float(ef.amount)
        for ef in ExamFee.objects.filter(
            school=school, school_class=student.school_class,
            academic_session=profile.current_academic_session,
        )
    }
    advance_available = _get_available_advance(student, profile.current_academic_session)
    return JsonResponse({
        'items': items,
        'due_month_options': _get_unpaid_month_options(student, school),
        'student_class': str(student.school_class),
        'session': profile.current_academic_session,
        'has_transport': bool(student.transport_opted and student.transport_amount),
        'advance_available': float(advance_available),
        'exam_amounts': exam_amounts,
    })


@school_only
def lump_sum_preview_ajax(request):
    """Return auto-distribution breakdown for a given student + lump-sum amount."""
    school = request.user.school
    student_id = request.GET.get('student_id')
    amount_str = request.GET.get('amount', '0').strip()

    if not student_id:
        return JsonResponse({'error': 'No student selected'}, status=400)

    try:
        cash_amount = Decimal(amount_str)
        if cash_amount <= 0:
            raise ValueError
    except (InvalidOperation, ValueError):
        return JsonResponse({'error': 'Invalid amount'}, status=400)

    student = get_object_or_404(Student, pk=student_id, school=school, status=Student.STATUS_ACTIVE)
    profile = SchoolProfile.get_for_school(school)
    session = profile.current_academic_session

    advance_available = _get_available_advance(student, session)
    result = _distribute_lump_sum(student, school, cash_amount, advance_available, profile)

    monthly_fee = _get_monthly_fee_total(student, school, session)
    transport_fee = _get_monthly_transport_fee(student)
    month_label_map = dict(MONTH_CHOICES)
    paid_tokens = set(result['paid_month_tokens'])
    session_months = get_session_months(profile.session_start_month, profile.session_end_month)

    months_breakdown = []
    for month in session_months:
        if month not in paid_tokens:
            continue
        transport_token = f'{month}_transport'
        has_transport_paid = transport_token in paid_tokens
        months_breakdown.append({
            'month': month_label_map.get(month, month.title()),
            'tuition': float(monthly_fee),
            'transport': float(transport_fee) if has_transport_paid else None,
            'row_total': float(monthly_fee + (transport_fee if has_transport_paid else Decimal('0'))),
        })

    return JsonResponse({
        'months_breakdown': months_breakdown,
        'exam_fee_items': result['exam_fee_items'],
        'gross_amount': float(result['gross_amount']),
        'advance_available': float(advance_available),
        'advance_balance': float(result['advance_balance']),
        'cash_amount': float(cash_amount),
        'nothing_to_pay': not months_breakdown and not result['exam_fee_items'],
    })


# ── Superuser Dashboard ────────────────────────────────────────────────────────

@super_only
def super_dashboard(request):
    su_settings = SuperUserSettings.get_solo()
    schools = School.objects.all()  # Include all schools (active and inactive)
    rows = []
    
    system_session = get_current_academic_session()
    profiles = {p.school_id: p for p in SchoolProfile.objects.all()}

    # Precompute student counts per (school, session) - only active students
    student_counts = {
        (r['school_id'], r['academic_session']): r['cnt']
        for r in Student.objects.filter(status=Student.STATUS_ACTIVE).values('school_id', 'academic_session').annotate(cnt=Count('id'))
    }

    for school in schools:
        profile = profiles.get(school.pk)
        current_session = profile.current_academic_session if profile else ''

        # Determine which sessions to include (matching super_school_fee_dashboard logic)
        sessions_to_show = {current_session} if current_session else set()

        # If this school was renewed from the system session, include the previous session
        try:
            start_year = int(current_session.split('-')[0])
            prev_session = format_academic_session(start_year - 1)
            if prev_session == system_session:
                sessions_to_show.add(prev_session)
        except (ValueError, IndexError, AttributeError):
            pass

        # Also include any sessions with billing records or SchoolSessionRecord entries
        billed_sessions = set(
            SchoolBillingPayment.objects.filter(school=school)
            .values_list('academic_session', flat=True)
            .distinct()
        )
        sessions_to_show.update(billed_sessions)

        # Add any historical sessions from SchoolSessionRecord
        session_record_sessions = set(
            SchoolSessionRecord.objects.filter(school=school)
            .values_list('academic_session', flat=True)
        )
        sessions_to_show.update(session_record_sessions)

        # Build rows for each session
        for session in sorted(sessions_to_show, reverse=True):
            is_current = session == current_session
            active_students = student_counts.get((school.pk, session), 0)
            rows.append({
                'school': school,
                'session': session,
                'is_current_session': is_current,
                # A past-session row is already "renewed" — hide the Renew button for it
                'is_renewed': not is_current,
                'student_count': active_students,
                'fee_per_student': school.fee_per_student,
                'billing_start_month': profile.billing_start_month if profile else '—',
                'billing_end_month': profile.billing_end_month if profile else '—',
            })

    unique_sessions = sorted(
        {r['session'] for r in rows if r['session'] != '—'},
        reverse=True,
    )
    
    # Calculate total students: sum from current session only to avoid double-counting
    # across schools with multiple sessions
    default_session = su_settings.default_session or system_session
    total_active_students = sum(
        r['student_count'] for r in rows 
        if r['session'] == default_session
    )
    
    return render(request, 'superuser/dashboard.html', {
        'rows': rows,
        'total_schools': School.objects.count(),
        'total_students': total_active_students,
        'active_schools': School.objects.filter(is_active=True).count(),
        'inactive_schools': School.objects.filter(is_active=False).count(),
        'session_options': unique_sessions,
        'system_session': system_session,
        'default_session': default_session,
    })


_MONTH_ABBR = {
    'april': 'Apr', 'may': 'May', 'june': 'Jun', 'july': 'Jul',
    'august': 'Aug', 'september': 'Sep', 'october': 'Oct', 'november': 'Nov',
    'december': 'Dec', 'january': 'Jan', 'february': 'Feb', 'march': 'Mar',
}


def _billed_months_and_label(session, b_start, b_end, is_current):
    """Return (month_count, period_label) for a school's billing row.

    Current session: count months from b_start up to today's month.
    Past sessions: count the full b_start → b_end range.
    """
    today = timezone.localdate()
    current_month_key = CAL_TO_MONTH[today.month]

    all_months = get_session_months(b_start, b_end)

    if is_current:
        if current_month_key in all_months:
            billed = all_months[:all_months.index(current_month_key) + 1]
        else:
            billed = []  # billing window not yet reached
    else:
        billed = all_months

    if not billed:
        return 0, '—'

    s_abbr = _MONTH_ABBR.get(billed[0], billed[0].capitalize())
    e_abbr = _MONTH_ABBR.get(billed[-1], billed[-1].capitalize())
    label = s_abbr if billed[0] == billed[-1] else f"{s_abbr} – {e_abbr}"
    return len(billed), label


@super_only
def super_school_fee_dashboard(request):
    su_settings = SuperUserSettings.get_solo()
    schools = School.objects.filter(is_active=True)
    rows = []
    total_monthly = Decimal('0.00')
    total_collected = Decimal('0.00')

    system_session = get_current_academic_session()
    profiles = {p.school_id: p for p in SchoolProfile.objects.all()}
    session_records = {}
    for sr in SchoolSessionRecord.objects.all():
        session_records[(sr.school_id, sr.academic_session)] = sr
    for school in schools:
        profile = profiles.get(school.pk)
        current_session = profile.current_academic_session if profile else ''
        active_students = Student.objects.filter(school=school, status=Student.STATUS_ACTIVE).count()
        per_month = school.fee_per_student * active_students

        # Start with the current session
        sessions_to_show = {current_session} if current_session else set()

        # If this school was renewed from the system session, always include that
        # previous session row — even when no billing was recorded yet
        try:
            start_year = int(current_session.split('-')[0])
            prev_session = format_academic_session(start_year - 1)
            if prev_session == system_session:
                sessions_to_show.add(prev_session)
        except (ValueError, IndexError, AttributeError):
            pass

        # Also surface any older sessions that already have billing records
        billed_sessions = set(
            SchoolBillingPayment.objects.filter(school=school)
            .values_list('academic_session', flat=True)
            .distinct()
        )
        sessions_to_show.update(billed_sessions)

        all_sessions = sorted(sessions_to_show, reverse=True)

        for session in all_sessions:
            is_current = session == current_session
            paid = SchoolBillingPayment.objects.filter(
                school=school, academic_session=session
            ).aggregate(t=Sum('amount_paid'))['t'] or Decimal('0.00')

            if is_current:
                # Always read from SchoolProfile for current session — it is
                # updated immediately when school details are edited.
                p = profile or SchoolProfile.get_for_school(school)
                b_start, b_end = p.billing_start_month, p.billing_end_month
            else:
                sr = session_records.get((school.pk, session))
                if sr:
                    b_start, b_end = sr.billing_start_month, sr.billing_end_month
                else:
                    p = profile or SchoolProfile.get_for_school(school)
                    b_start, b_end = p.billing_start_month, p.billing_end_month

            month_count, billing_period = _billed_months_and_label(session, b_start, b_end, is_current)
            total_due = per_month * month_count
            balance = max(total_due - paid, Decimal('0.00'))

            rows.append({
                'school': school,
                'session': session,
                'is_current_session': is_current,
                'active_students': active_students,
                'fee_per_student': school.fee_per_student,
                'monthly_bill': per_month,
                'total_due': total_due,
                'month_count': month_count,
                'total_paid': paid,
                'balance': balance,
                'is_paid_up': balance <= Decimal('0.00'),
                'billing_period': billing_period,
            })

            # KPI totals: current session only to avoid double-counting
            if is_current:
                total_monthly += total_due
                total_collected += paid

    total_pending = max(total_monthly - total_collected, Decimal('0.00'))
    session_options = sorted(
        {r['session'] for r in rows if r.get('session')},
        reverse=True,
    )
    return render(request, 'superuser/school_fee_dashboard.html', {
        'rows': rows,
        'total_monthly': total_monthly,
        'total_collected': total_collected,
        'total_pending': total_pending,
        'session_options': session_options,
        'default_session': su_settings.default_session or system_session,
    })


@super_only
def super_collect_fee(request, school_pk):
    school = get_object_or_404(School, pk=school_pk)
    num_students = Student.objects.filter(school=school, status=Student.STATUS_ACTIVE).count()
    fee_per_student = school.fee_per_student

    profile = SchoolProfile.get_for_school(school)
    requested_session = request.GET.get('session', '').strip()
    current_session = requested_session if requested_session else profile.current_academic_session
    all_billing_months = _get_school_billing_months(school)  # [(token, label), …]

    # Track per-month student coverage: regular payments set the baseline,
    # adjustment payments add additional students on top.
    regular_month_students = {}   # {token: num_students at time of regular payment}
    adjustment_month_students = {}  # {token: additional students from adjustments}
    all_session_payments = SchoolBillingPayment.objects.filter(school=school, academic_session=current_session)
    for bp in all_session_payments:
        for m in (bp.payment_months or []):
            m = str(m)
            if bp.is_adjustment:
                adjustment_month_students[m] = adjustment_month_students.get(m, 0) + bp.num_students
            else:
                # Take the most recent regular payment value (last write wins)
                regular_month_students[m] = regular_month_students.get(m, 0) + bp.num_students

    paid_months = set(regular_month_students.keys())
    unpaid_months = [(val, label) for val, label in all_billing_months if val not in paid_months]

    # Months that were paid but student count has since grown → adjustment due
    adjustment_rows = []
    total_adjustment_due = Decimal('0.00')
    for token, label in all_billing_months:
        if token in paid_months:
            covered = regular_month_students.get(token, 0) + adjustment_month_students.get(token, 0)
            deficit = num_students - covered
            if deficit > 0:
                adj_amt = fee_per_student * deficit
                adjustment_rows.append({
                    'token': token,
                    'label': label,
                    'covered': covered,
                    'deficit': deficit,
                    'amount': adj_amt,
                })
                total_adjustment_due += adj_amt

    if request.method == 'POST':
        payment_date = request.POST.get('payment_date')
        payment_type = request.POST.get('payment_type', 'regular')

        if payment_type == 'adjustment':
            adj_by_token = {r['token']: r for r in adjustment_rows}
            selected_months = [m for m in request.POST.getlist('adjustment_months') if m in adj_by_token]
            if payment_date and selected_months:
                amount_paid = sum(adj_by_token[m]['amount'] for m in selected_months)
                # num_students = delta (additional students being covered)
                delta = max(adj_by_token[m]['deficit'] for m in selected_months)
                month_labels = ', '.join(adj_by_token[m]['label'] for m in selected_months)
                SchoolBillingPayment.objects.create(
                    school=school,
                    academic_session=current_session,
                    payment_date=payment_date,
                    num_students=delta,
                    fee_per_student=fee_per_student,
                    payment_months=selected_months,
                    amount_paid=amount_paid,
                    is_adjustment=True,
                    note=f'Adjustment for {delta} additional student(s): {month_labels}',
                )
                messages.success(request, f'Adjustment of ₹{amount_paid} recorded for {school.name}.')
                return redirect('super_school_fee_dashboard')
        else:
            valid_tokens = {val for val, _ in all_billing_months}
            selected_months = [m for m in request.POST.getlist('payment_months')
                               if m in valid_tokens and m not in paid_months]
            if payment_date and selected_months:
                amount_paid = fee_per_student * num_students * len(selected_months)
                SchoolBillingPayment.objects.create(
                    school=school,
                    academic_session=current_session,
                    payment_date=payment_date,
                    num_students=num_students,
                    fee_per_student=fee_per_student,
                    payment_months=selected_months,
                    amount_paid=amount_paid,
                )
                messages.success(request, f'Payment of ₹{amount_paid} recorded for {school.name} ({current_session}).')
                return redirect('super_school_fee_dashboard')

    billing_history = all_session_payments.order_by('-payment_date')
    total_billed = billing_history.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0.00')
    total_months = len(all_billing_months)
    paid_count = total_months - len(unpaid_months)

    return render(request, 'superuser/collect_fee.html', {
        'school': school,
        'num_students': num_students,
        'fee_per_student': fee_per_student,
        'per_month_amount': fee_per_student * num_students,
        'unpaid_months': unpaid_months,
        'adjustment_rows': adjustment_rows,
        'total_adjustment_due': total_adjustment_due,
        'billing_history': billing_history,
        'total_billed': total_billed,
        'total_months': total_months,
        'paid_count': paid_count,
        'current_session': current_session,
        'billing_start': all_billing_months[0][1] if all_billing_months else '',
        'billing_end': all_billing_months[-1][1] if all_billing_months else '',
    })


@super_only
def super_promote_school(request, pk):
    if request.method == 'POST':
        school = get_object_or_404(School, pk=pk)
        profile = SchoolProfile.get_for_school(school)
        current_session = profile.current_academic_session
        try:
            start_year = int(current_session.split('-')[0])
        except (ValueError, IndexError):
            start_year = get_session_start_year(timezone.localdate(), profile.session_start_month)
        next_session = format_academic_session(start_year + 1)

        if profile.current_academic_session == next_session:
            messages.warning(request, f'{school.name} is already on session {next_session}.')
            return redirect('super_dashboard')

        SchoolSessionRecord.objects.update_or_create(
            school=school,
            academic_session=next_session,
            defaults={
                'session_start_month': profile.session_start_month,
                'session_end_month': profile.session_end_month,
                'billing_start_month': profile.billing_start_month,
                'billing_end_month': profile.billing_end_month,
            },
        )

        existing_structures = FeeStructure.objects.filter(
            school_class__school=school,
            academic_session=current_session,
        )
        new_structures = []
        for fs in existing_structures:
            already_exists = FeeStructure.objects.filter(
                school_class=fs.school_class,
                fee_category=fs.fee_category,
                academic_session=next_session,
            ).exists()
            if not already_exists:
                new_structures.append(FeeStructure(
                    school_class=fs.school_class,
                    fee_category=fs.fee_category,
                    academic_session=next_session,
                    amount=fs.amount,
                    frequency=fs.frequency,
                ))
        if new_structures:
            FeeStructure.objects.bulk_create(new_structures)

        profile.current_academic_session = next_session
        profile.save()

        messages.success(
            request,
            f'{school.name} renewed to {next_session}. '
            f'Fee structures copied; previous session data is preserved.'
        )
    return redirect('super_dashboard')


@super_only
def super_settings(request):
    settings_obj = SuperUserSettings.get_solo()
    if request.method == 'POST':
        form = SuperUserSettingsForm(request.POST, request.FILES, instance=settings_obj)
        if form.is_valid():
            instance = form.save(commit=False)
            if not request.FILES.get('logo'):
                instance.logo = settings_obj.logo
            # Save default_session directly from POST — it's managed outside the form
            raw_session = request.POST.get('default_session', '').strip()
            instance.default_session = raw_session
            instance.save()
            # Refresh so the render below reflects the just-saved value
            settings_obj = SuperUserSettings.get_solo()
            messages.success(request, 'Settings saved successfully.')
            return redirect('super_settings')
        else:
            messages.error(request, 'Could not save settings. Please check the highlighted fields.')
    else:
        form = SuperUserSettingsForm(instance=settings_obj)
    session_choices = get_academic_session_choices(past_years=5, future_years=10)
    # Ensure the currently saved session is always in the list
    saved_session = settings_obj.default_session or ''
    if saved_session and saved_session not in [v for v, _ in session_choices]:
        session_choices.insert(0, (saved_session, saved_session))
    return render(request, 'superuser/settings.html', {
        'form': form,
        'settings_obj': settings_obj,
        'session_choices': session_choices,
        'selected_session': saved_session,
    })


# ── Configuration ─────────────────────────────────────────────────────────────

@school_only
def config_view(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    classes = SchoolClass.objects.filter(school=school).prefetch_related('fee_structures__fee_category')
    session_choices = get_academic_session_choices(past_years=2, future_years=10)

    if request.method == 'POST':
        if 'save_profile' in request.POST:
            session = request.POST.get('current_academic_session', '').strip()
            start_month = request.POST.get('session_start_month', '').strip()
            end_month = request.POST.get('session_end_month', '').strip()
            valid_months = {m for m, _ in MONTH_CHOICES}
            if not session:
                messages.error(request, 'Session is required.')
            elif start_month not in valid_months or end_month not in valid_months:
                messages.error(request, 'Invalid session month selection.')
            else:
                profile.current_academic_session = session
                profile.session_start_month = start_month
                profile.session_end_month = end_month
                profile.save(update_fields=['current_academic_session', 'session_start_month', 'session_end_month'])
                SchoolSessionRecord.objects.update_or_create(
                    school=school,
                    academic_session=session,
                    defaults={
                        'session_start_month': start_month,
                        'session_end_month': end_month,
                        'billing_start_month': profile.billing_start_month,
                        'billing_end_month': profile.billing_end_month,
                    },
                )
                messages.success(request, 'School profile updated.')

        elif 'add_class' in request.POST:
            class_name = request.POST.get('class_name', '').strip()
            try:
                amount = Decimal(request.POST.get('monthly_fee', ''))
                assert amount >= 0
            except (InvalidOperation, AssertionError):
                amount = None
            if not class_name or amount is None:
                messages.error(request, 'Enter valid class name and fee.')
            else:
                klass = SchoolClass.objects.create(school=school, name=class_name)
                category, _ = FeeCategory.objects.get_or_create(school=school, name='Monthly Fee', defaults={'is_active': True})
                FeeStructure.objects.update_or_create(
                    school_class=klass, fee_category=category, academic_session=profile.current_academic_session,
                    defaults={'amount': amount, 'frequency': FeeStructure.FREQUENCY_MONTHLY},
                )
                messages.success(request, 'Class added.')

        elif 'update_class' in request.POST:
            class_id = request.POST.get('class_id')
            class_name = request.POST.get('class_name', '').strip()
            try:
                amount = Decimal(request.POST.get('monthly_fee', ''))
                assert amount >= 0
            except (InvalidOperation, AssertionError):
                amount = None
            if not class_id or not class_name or amount is None:
                messages.error(request, 'Enter valid class name and fee.')
            else:
                klass = get_object_or_404(SchoolClass, pk=class_id, school=school)
                klass.name = class_name
                klass.save(update_fields=['name'])
                category, _ = FeeCategory.objects.get_or_create(school=school, name='Monthly Fee', defaults={'is_active': True})
                FeeStructure.objects.update_or_create(
                    school_class=klass, fee_category=category, academic_session=profile.current_academic_session,
                    defaults={'amount': amount, 'frequency': FeeStructure.FREQUENCY_MONTHLY},
                )
                messages.success(request, 'Class updated.')

        elif 'delete_class' in request.POST:
            klass = get_object_or_404(SchoolClass, pk=request.POST.get('class_id'), school=school)
            student_count = Student.objects.filter(school_class=klass).count()
            Student.objects.filter(school_class=klass).delete()
            klass.delete()
            messages.success(request, f'Class "{klass.name}" and {student_count} student(s) deleted.')

        elif 'add_exam_fee' in request.POST:
            class_id = request.POST.get('exam_class_id')
            exam_name = request.POST.get('exam_name', '').strip()
            session = profile.current_academic_session
            try:
                amount = Decimal(request.POST.get('exam_amount', ''))
                assert amount >= 0
            except (InvalidOperation, AssertionError):
                amount = None
            if not class_id or not exam_name or amount is None:
                messages.error(request, 'All exam fee fields are required.')
            else:
                klass = get_object_or_404(SchoolClass, pk=class_id, school=school)
                ExamFee.objects.create(school=school, school_class=klass, exam_name=exam_name, academic_session=session, amount=amount)
                messages.success(request, 'Exam fee added.')

        elif 'update_exam_fee' in request.POST:
            exam_id = request.POST.get('exam_fee_id')
            exam_name = request.POST.get('exam_name', '').strip()
            session = request.POST.get('exam_session', '').strip()
            try:
                amount = Decimal(request.POST.get('exam_amount', ''))
                assert amount >= 0
            except (InvalidOperation, AssertionError):
                amount = None
            class_id = request.POST.get('exam_class_id')
            if not exam_id or not exam_name or not session or amount is None or not class_id:
                messages.error(request, 'All exam fee fields are required.')
            else:
                exam = get_object_or_404(ExamFee, pk=exam_id, school=school)
                klass = get_object_or_404(SchoolClass, pk=class_id, school=school)
                exam.exam_name = exam_name
                exam.academic_session = session
                exam.amount = amount
                exam.school_class = klass
                exam.save()
                messages.success(request, 'Exam fee updated.')

        elif 'delete_exam_fee' in request.POST:
            exam = get_object_or_404(ExamFee, pk=request.POST.get('exam_fee_id'), school=school)
            exam.delete()
            messages.success(request, 'Exam fee deleted.')

        elif 'save_wa_config' in request.POST:
            wa_config, _ = WhatsAppConfig.objects.get_or_create(school=school)
            wa_config.phone_number_id                = request.POST.get('phone_number_id', '').strip()
            wa_config.waba_id                        = request.POST.get('waba_id', '').strip()
            wa_config.access_token                   = request.POST.get('access_token', '').strip()
            wa_config.template_name                  = request.POST.get('template_name', '').strip()
            wa_config.template_language              = request.POST.get('template_language', 'en').strip()
            wa_config.announcement_template_name     = request.POST.get('announcement_template_name', '').strip()
            wa_config.announcement_template_language = request.POST.get('announcement_template_language', 'en').strip()
            wa_config.is_active                      = bool(request.POST.get('is_active'))
            wa_config.save()
            messages.success(request, 'WhatsApp configuration saved.')

        return redirect('config_view')

    wa_config, _ = WhatsAppConfig.objects.get_or_create(school=school)
    exam_fees = ExamFee.objects.filter(school=school).select_related('school_class')
    return render(request, 'school/config/config.html', {
        'object': profile,
        'school': school,
        'classes': classes,
        'session_choices': session_choices,
        'month_choices': MONTH_CHOICES,
        'exam_fees': exam_fees,
        'wa_config': wa_config,
    })


# ─────────────────────────────────────────────────────────────────────────────
# WhatsApp Reminder Views
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@school_only
def whatsapp_dashboard(request):
    """Show WhatsApp config + list of students with pending fees."""
    school = request.user.school
    profile = SchoolProfile.objects.filter(school=school).first()
    wa_config, _ = WhatsAppConfig.objects.get_or_create(school=school)

    # ── Build pending-fee student list ────────────────────────────────
    current_session = profile.current_academic_session if profile else ''
    students = Student.objects.filter(
        school=school, status=Student.STATUS_ACTIVE
    ).select_related('school_class').order_by('school_class', 'name')

    pending_rows = []
    for student in students:
        paid = FeePayment.objects.filter(
            student=student, academic_session=current_session
        ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')

        # total due so far (all active fee structures * months elapsed)
        structures = FeeStructure.objects.filter(
            fee_category__school=school,
            fee_category__is_active=True,
            academic_session=current_session,
            school_class=student.school_class,
        )
        monthly_fee = sum(s.amount for s in structures if s.frequency == 'monthly')
        if student.transport_opted and student.transport_amount:
            monthly_fee += student.transport_amount

        # months elapsed in session so far
        session_start = profile.session_start_month if profile else 'april'
        today = date.today()
        start_cal = MONTH_TO_CAL.get(session_start, 4)
        session_year = int(current_session[:4]) if current_session else today.year
        start_date = date(session_year, start_cal, 1)
        months_elapsed = max(1, (today.year - start_date.year) * 12 + today.month - start_date.month + 1)

        total_due  = monthly_fee * months_elapsed
        balance    = total_due - paid

        if balance > 0:
            pending_rows.append({
                'student': student,
                'paid':    paid,
                'due':     total_due,
                'balance': balance,
                'phone':   student.father_phone,
            })

    classes = SchoolClass.objects.filter(school=school).order_by('name')
    session_options = sorted({
        s.academic_session for s in Student.objects.filter(school=school, status=Student.STATUS_ACTIVE)
    })
    return render(request, 'school/whatsapp/dashboard.html', {
        'wa_config':       wa_config,
        'pending_rows':    pending_rows,
        'school':          school,
        'current_session': current_session,
        'classes':         classes,
        'session_options': session_options,
    })


@login_required
@school_only
def whatsapp_send(request):
    """AJAX endpoint — send a WhatsApp reminder to one parent."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)

    school     = request.user.school
    wa_config  = get_object_or_404(WhatsAppConfig, school=school)
    student_pk = request.POST.get('student_pk')
    student    = get_object_or_404(Student, pk=student_pk, school=school)
    balance    = request.POST.get('balance', '0')

    if not wa_config.is_active:
        return JsonResponse({'ok': False, 'error': 'WhatsApp is not enabled. Please configure and enable it first.'})

    if not wa_config.phone_number_id or not wa_config.access_token:
        return JsonResponse({'ok': False, 'error': 'Missing Phone Number ID or Access Token in settings.'})

    phone = student.father_phone.strip()
    if not phone:
        return JsonResponse({'ok': False, 'error': 'No phone number on record for this student.'})

    # Normalize to E.164 (assume Indian number if 10 digits)
    phone = phone.lstrip('+').replace(' ', '').replace('-', '')
    if len(phone) == 10 and phone.isdigit():
        phone = '91' + phone

    template_name = wa_config.template_name.strip()
    if not template_name:
        return JsonResponse({'ok': False, 'error': 'Template 1 (fee reminder) not set. Add it in Configuration → WhatsApp.'})

    template_language = (wa_config.template_language or '').strip() or 'en'

    template_payload = {
        "name": template_name,
        "language": {"code": template_language},
        "components": [
            {
                "type": "body",
                "parameters": [
                    {"type": "text", "text": student.name},
                    {"type": "text", "text": student.roll_number},
                    {"type": "text", "text": f"Rs.{balance}"},
                    {"type": "text", "text": school.name},
                ]
            }
        ]
    }

    payload = {
        "messaging_product": "whatsapp",
        "to": phone,
        "type": "template",
        "template": template_payload,
    }

    url = f"https://graph.facebook.com/v25.0/{wa_config.phone_number_id}/messages"
    data = json.dumps(payload).encode('utf-8')
    req  = urllib.request.Request(
        url, data=data,
        headers={
            'Authorization': f'Bearer {wa_config.access_token}',
            'Content-Type':  'application/json',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read())
        return JsonResponse({'ok': True, 'message_id': result.get('messages', [{}])[0].get('id', '')})
    except urllib.error.HTTPError as e:
        err_body = e.read().decode('utf-8', errors='replace')
        try:
            err_msg = json.loads(err_body).get('error', {}).get('message', err_body)
        except Exception:
            err_msg = err_body
        return JsonResponse({'ok': False, 'error': err_msg})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required
@school_only
def announcement_dashboard(request):
    """Dedicated Announcement page — send a custom message to any/all students."""
    school    = request.user.school
    wa_config, _ = WhatsAppConfig.objects.get_or_create(school=school)

    all_students = Student.objects.filter(
        school=school, status=Student.STATUS_ACTIVE
    ).select_related('school_class').order_by('school_class', 'name')

    classes = SchoolClass.objects.filter(school=school).order_by('name')
    session_options = sorted({s.academic_session for s in all_students})

    return render(request, 'school/whatsapp/announcement.html', {
        'wa_config':       wa_config,
        'all_students':    all_students,
        'classes':         classes,
        'session_options': session_options,
    })


@login_required
@school_only
def whatsapp_templates_debug(request):
    """Return the exact template names and language codes stored in Meta for this WABA."""
    school    = request.user.school
    wa_config = get_object_or_404(WhatsAppConfig, school=school)

    if not wa_config.access_token:
        return JsonResponse({'ok': False, 'error': 'Access Token not configured.'})
    if not wa_config.waba_id:
        return JsonResponse({'ok': False, 'error': 'WhatsApp Business Account ID (WABA ID) not configured. Add it in Configuration → WhatsApp.'})

    headers = {'Authorization': f'Bearer {wa_config.access_token}'}

    try:
        url = f"https://graph.facebook.com/v25.0/{wa_config.waba_id}/message_templates?fields=name,language,status"
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=10) as r:
            tpl_data = json.loads(r.read())
        templates = [
            {'name': t['name'], 'language': t['language'], 'status': t.get('status')}
            for t in tpl_data.get('data', [])
        ]
        return JsonResponse({'ok': True, 'templates': templates})
    except urllib.error.HTTPError as e:
        err_body = e.read().decode('utf-8', errors='replace')
        try:
            err_msg = json.loads(err_body).get('error', {}).get('message', err_body)
        except Exception:
            err_msg = err_body
        return JsonResponse({'ok': False, 'error': err_msg})


@login_required
@school_only
def whatsapp_announce(request):
    """AJAX endpoint — send a custom announcement via WhatsApp to one parent."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)

    school     = request.user.school
    wa_config  = get_object_or_404(WhatsAppConfig, school=school)
    student_pk = request.POST.get('student_pk')
    message    = request.POST.get('message', '').strip()
    student    = get_object_or_404(Student, pk=student_pk, school=school)

    if not wa_config.is_active:
        return JsonResponse({'ok': False, 'error': 'WhatsApp is not enabled. Enable it in Configuration.'})
    if not wa_config.phone_number_id or not wa_config.access_token:
        return JsonResponse({'ok': False, 'error': 'Missing Phone Number ID or Access Token in Configuration.'})
    ann_template_name = wa_config.announcement_template_name.strip()
    if not ann_template_name:
        return JsonResponse({'ok': False, 'error': 'Template 2 (announcement) not set. Add it in Configuration → WhatsApp.'})
    if not message:
        return JsonResponse({'ok': False, 'error': 'Message cannot be empty.'})

    phone = (student.father_phone or '').strip()
    if not phone:
        return JsonResponse({'ok': False, 'error': 'No phone number on record for this student.'})

    # Normalize to E.164 (assume Indian number if 10 digits)
    phone = phone.lstrip('+').replace(' ', '').replace('-', '')
    if len(phone) == 10 and phone.isdigit():
        phone = '91' + phone

    ann_template_language = (wa_config.announcement_template_language or '').strip() or 'en'

    template_payload = {
        "name": ann_template_name,
        "language": {"code": ann_template_language},
        "components": [
            {
                "type": "body",
                "parameters": [
                    {"type": "text", "text": message},
                ]
            }
        ]
    }

    payload = {
        "messaging_product": "whatsapp",
        "to": phone,
        "type": "template",
        "template": template_payload,
    }

    url  = f"https://graph.facebook.com/v25.0/{wa_config.phone_number_id}/messages"
    data = json.dumps(payload).encode('utf-8')
    req  = urllib.request.Request(
        url, data=data,
        headers={
            'Authorization': f'Bearer {wa_config.access_token}',
            'Content-Type':  'application/json',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read())
        return JsonResponse({'ok': True, 'message_id': result.get('messages', [{}])[0].get('id', '')})
    except urllib.error.HTTPError as e:
        err_body = e.read().decode('utf-8', errors='replace')
        try:
            err_msg = json.loads(err_body).get('error', {}).get('message', err_body)
        except Exception:
            err_msg = err_body
        return JsonResponse({'ok': False, 'error': err_msg})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


# ── Teachers ──────────────────────────────────────────────────────────────────

@school_only
def teacher_list(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    session_choices = get_academic_session_choices(past_years=2, future_years=10)

    selected_session = request.GET.get('session', profile.current_academic_session)
    q = request.GET.get('q', '').strip()
    class_id = request.GET.get('class', '')
    status_filter = request.GET.get('status', '')
    employment_filter = request.GET.get('employment', '')

    teachers = Teacher.objects.filter(school=school).select_related('class_teacher_of')

    if selected_session:
        teachers = teachers.filter(academic_session=selected_session)
    if q:
        teachers = teachers.filter(
            Q(name__icontains=q) | Q(employee_id__icontains=q) |
            Q(phone__icontains=q) | Q(subjects_taught__icontains=q)
        )
    if class_id:
        teachers = teachers.filter(class_teacher_of_id=class_id)
    if status_filter:
        teachers = teachers.filter(status=status_filter)
    if employment_filter:
        teachers = teachers.filter(employment_type=employment_filter)

    total_active = Teacher.objects.filter(school=school, status=Teacher.STATUS_ACTIVE, academic_session=selected_session).count()
    total_inactive = Teacher.objects.filter(school=school, status__in=[Teacher.STATUS_INACTIVE, Teacher.STATUS_RESIGNED], academic_session=selected_session).count()

    paginator = Paginator(teachers.order_by('name'), 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'school/teachers/teacher_list.html', {
        'teachers': page_obj,
        'page_obj': page_obj,
        'is_paginated': paginator.num_pages > 1,
        'classes': SchoolClass.objects.filter(school=school),
        'session_choices': session_choices,
        'selected_session': selected_session,
        'selected_class': class_id,
        'selected_q': q,
        'selected_status': status_filter,
        'selected_employment': employment_filter,
        'total_active': total_active,
        'total_inactive': total_inactive,
        'total_filtered': teachers.count(),
        'status_choices': Teacher.STATUS_CHOICES,
        'employment_choices': Teacher.EMPLOYMENT_CHOICES,
    })


@school_only
def teacher_create(request):
    school = request.user.school
    if request.method == 'POST':
        form = TeacherForm(request.POST, school=school)
        if form.is_valid():
            teacher = form.save(commit=False)
            teacher.school = school
            teacher.save()
            messages.success(request, f'Teacher {teacher.name} ({teacher.employee_id}) added successfully.')
            return redirect('teacher_list')
    else:
        form = TeacherForm(school=school)
    return render(request, 'school/teachers/teacher_form.html', {
        'form': form,
        'title': 'Add Teacher',
        'submit_label': 'Add Teacher',
    })


@school_only
def teacher_edit(request, pk):
    school = request.user.school
    teacher = get_object_or_404(Teacher, pk=pk, school=school)
    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=teacher, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f'{teacher.name} updated successfully.')
            return redirect('teacher_list')
    else:
        form = TeacherForm(instance=teacher, school=school)
    return render(request, 'school/teachers/teacher_form.html', {
        'form': form,
        'teacher': teacher,
        'title': f'Edit — {teacher.name}',
        'submit_label': 'Save Changes',
    })


@school_only
def teacher_delete(request, pk):
    school = request.user.school
    teacher = get_object_or_404(Teacher, pk=pk, school=school)
    if request.method == 'POST':
        name = teacher.name
        teacher.delete()
        messages.success(request, f'Teacher {name} deleted.')
        return redirect('teacher_list')
    return render(request, 'school/teachers/teacher_confirm_delete.html', {'teacher': teacher})


@school_only
def teacher_promote(request, pk):
    if request.method != 'POST':
        return redirect('teacher_list')
    school = request.user.school
    teacher = get_object_or_404(Teacher, pk=pk, school=school, status=Teacher.STATUS_ACTIVE)

    try:
        pct = float(request.POST.get('increment_pct', ''))
        if pct <= 0 or pct > 100:
            raise ValueError
    except (ValueError, TypeError):
        messages.error(request, 'Invalid increment percentage.')
        return redirect('teacher_list')

    old_salary = teacher.monthly_salary
    old_session = teacher.academic_session

    increment = (old_salary * Decimal(str(pct)) / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    teacher.monthly_salary = old_salary + increment

    try:
        start_year = int(old_session.split('-')[0])
        teacher.academic_session = f"{start_year + 1}-{(start_year + 2) % 100:02d}"
    except (ValueError, IndexError, AttributeError):
        teacher.academic_session = get_current_academic_session(timezone.localdate())

    teacher.save(update_fields=['monthly_salary', 'academic_session', 'updated_at'])
    messages.success(
        request,
        f'{teacher.name} promoted — session {old_session} → {teacher.academic_session}, salary ₹{old_salary} → ₹{teacher.monthly_salary} (+{pct}%).',
    )
    return redirect('teacher_list')


# ── Salary Submission ──────────────────────────────────────────────────────────

def _compute_salary_total(teacher, months, session):
    """Compute total salary for selected months, prorating the joining month."""
    total = Decimal('0.00')
    session_start_year = int(session[:4]) if session and len(session) >= 4 else date.today().year
    joining_date = teacher.joining_date
    joining_month_start = date(joining_date.year, joining_date.month, 1)

    for m in months:
        cal_m = MONTH_TO_CAL.get(m, 1)
        m_year = session_start_year if cal_m >= 4 else session_start_year + 1
        month_start = date(m_year, cal_m, 1)
        if month_start == joining_month_start:
            total_days = calendar.monthrange(m_year, cal_m)[1]
            worked_days = total_days - joining_date.day + 1
            ratio = Decimal(worked_days) / Decimal(total_days)
            total += (teacher.monthly_salary * ratio).quantize(Decimal('0.01'))
        else:
            total += teacher.monthly_salary or Decimal('0.00')
    return total


def salary_dashboard(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    session_choices = get_academic_session_choices(past_years=2, future_years=10)

    selected_session = request.GET.get('session', profile.current_academic_session)
    q = request.GET.get('q', '').strip()
    pay_status = request.GET.get('pay_status', '')

    teachers = Teacher.objects.filter(school=school, academic_session=selected_session).order_by('name')
    if q:
        teachers = teachers.filter(Q(name__icontains=q) | Q(employee_id__icontains=q))

    # Paid months per teacher for selected session
    salary_records = SalaryPayment.objects.filter(
        school=school, academic_session=selected_session
    ).values('teacher_id', 'payment_months', 'amount_paid')

    paid_months_by_teacher = {}
    total_paid_by_teacher = {}
    for rec in salary_records:
        tid = rec['teacher_id']
        paid_months_by_teacher.setdefault(tid, set())
        paid_months_by_teacher[tid].update(rec['payment_months'] or [])
        total_paid_by_teacher[tid] = total_paid_by_teacher.get(tid, Decimal('0.00')) + rec['amount_paid']

    session_start_year = int(selected_session[:4]) if selected_session and len(selected_session) >= 4 else date.today().year
    today = timezone.localdate()
    profile_session = profile.current_academic_session

    # Cutoff date: salary is due only up to today for current session,
    # up to session end for past sessions, nothing for future sessions.
    if selected_session < profile_session:
        end_m = profile.session_end_month
        end_cal = MONTH_TO_CAL[end_m]
        end_year = session_start_year if end_cal >= 4 else session_start_year + 1
        cutoff_date = date(end_year, end_cal, 1)
    elif selected_session == profile_session:
        cutoff_date = date(today.year, today.month, 1)
    else:
        cutoff_date = None  # future session — nothing due

    rows = []
    for t in teachers:
        paid_months = paid_months_by_teacher.get(t.id, set())
        monthly_sal = t.monthly_salary or Decimal('0.00')

        # Salary is due from the teacher's joining month up to cutoff_date only
        joining_date = t.joining_date
        joining_month_start = date(joining_date.year, joining_date.month, 1)
        total_salary_due = Decimal('0.00')
        payable_count = 0

        if cutoff_date is not None:
            for v, _ in MONTH_CHOICES:
                cal_m = MONTH_TO_CAL.get(v, 1)
                m_year = session_start_year if cal_m >= 4 else session_start_year + 1
                month_start = date(m_year, cal_m, 1)
                if month_start < joining_month_start:
                    continue  # before this teacher joined
                if month_start > cutoff_date:
                    continue  # not yet due
                payable_count += 1
                if month_start == joining_month_start:
                    total_days = calendar.monthrange(m_year, cal_m)[1]
                    worked_days = total_days - joining_date.day + 1
                    total_salary_due += (monthly_sal * Decimal(worked_days) / Decimal(total_days)).quantize(Decimal('0.01'))
                else:
                    total_salary_due += monthly_sal

        paid_count = len(paid_months)
        pending_count = max(payable_count - paid_count, 0)
        total_paid = total_paid_by_teacher.get(t.id, Decimal('0.00'))
        balance = max(total_salary_due - total_paid, Decimal('0.00'))
        rows.append({
            'teacher': t,
            'paid_months': paid_months,
            'paid_count': paid_count,
            'payable_count': payable_count,
            'pending_count': pending_count,
            'total_salary_due': total_salary_due,
            'total_paid': total_paid,
            'balance': balance,
            'is_paid_up': balance <= Decimal('0.00'),
        })

    if pay_status == 'paid':
        rows = [r for r in rows if r['is_paid_up']]
    elif pay_status == 'pending':
        rows = [r for r in rows if not r['is_paid_up']]

    total_teachers = len(rows)
    total_due_all = sum(r['total_salary_due'] for r in rows)
    total_paid_all = sum(r['total_paid'] for r in rows)
    total_pending_all = max(total_due_all - total_paid_all, Decimal('0.00'))

    paginator = Paginator(rows, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'school/salary/salary_dashboard.html', {
        'rows': page_obj,
        'page_obj': page_obj,
        'is_paginated': paginator.num_pages > 1,
        'session_choices': session_choices,
        'selected_session': selected_session,
        'selected_q': q,
        'selected_pay_status': pay_status,
        'total_teachers': total_teachers,
        'total_due_all': total_due_all,
        'total_paid_all': total_paid_all,
        'total_pending_all': total_pending_all,
    })


@school_only
def salary_pay(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    teacher_id = request.GET.get('teacher_id') or request.POST.get('teacher')

    if request.method == 'POST':
        form = SalaryPaymentForm(request.POST, school=school)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.school = school
            payment.paid_by = request.user
            months = form.cleaned_data['payment_months']
            teacher = form.cleaned_data['teacher']
            session_val = form.cleaned_data['academic_session']
            payment.amount_paid = _compute_salary_total(teacher, months, session_val)
            payment.save()
            messages.success(request, f'Salary paid for {teacher.name} — {payment.get_payment_months_display()}.')
            return redirect('salary_dashboard')
    else:
        form = SalaryPaymentForm(school=school)
        if teacher_id:
            try:
                form.fields['teacher'].initial = int(teacher_id)
            except (ValueError, TypeError):
                pass

    # Determine session and cutoff date
    session = request.GET.get('session', profile.current_academic_session)
    session_start_year = int(session[:4]) if session and len(session) >= 4 else date.today().year
    today = timezone.localdate()
    profile_session = profile.current_academic_session

    if session < profile_session:
        end_cal = MONTH_TO_CAL[profile.session_end_month]
        end_year = session_start_year if end_cal >= 4 else session_start_year + 1
        cutoff_date = date(end_year, end_cal, 1)
    elif session == profile_session:
        cutoff_date = date(today.year, today.month, 1)
    else:
        cutoff_date = None  # future session — nothing payable

    # Build available month choices limited to cutoff (used when no teacher is selected)
    available_month_choices = []
    if cutoff_date is not None:
        for v, l in MONTH_CHOICES:
            cal_m = MONTH_TO_CAL.get(v, 1)
            m_year = session_start_year if cal_m >= 4 else session_start_year + 1
            if date(m_year, cal_m, 1) <= cutoff_date:
                available_month_choices.append((v, l))

    # Unpaid months for the selected teacher (from joining date up to cutoff only)
    unpaid_months = None
    selected_teacher = None
    month_amounts = {}
    if teacher_id:
        try:
            selected_teacher = Teacher.objects.get(pk=teacher_id, school=school)
            paid_months_qs = SalaryPayment.objects.filter(
                school=school, teacher=selected_teacher, academic_session=session,
            ).values_list('payment_months', flat=True)
            paid_set = set()
            for ml in paid_months_qs:
                paid_set.update(ml or [])

            unpaid_months = []
            joining_date = selected_teacher.joining_date
            joining_month_start = date(joining_date.year, joining_date.month, 1)

            for v, l in MONTH_CHOICES:
                if v in paid_set:
                    continue
                cal_m = MONTH_TO_CAL.get(v, 1)
                m_year = session_start_year if cal_m >= 4 else session_start_year + 1
                month_start = date(m_year, cal_m, 1)
                if month_start < joining_month_start:
                    continue  # before joining
                if cutoff_date is None or month_start > cutoff_date:
                    continue  # future month, not yet due
                if month_start == joining_month_start:
                    total_days = calendar.monthrange(m_year, cal_m)[1]
                    worked_days = total_days - joining_date.day + 1
                    ratio = Decimal(worked_days) / Decimal(total_days)
                    amount = (selected_teacher.monthly_salary * ratio).quantize(Decimal('0.01'))
                    unpaid_months.append((v, f'{l} ({worked_days}/{total_days} days — prorated)', amount, round(float(ratio), 6)))
                    month_amounts[v] = float(amount)
                else:
                    amount = (selected_teacher.monthly_salary or Decimal('0')).quantize(Decimal('0.01'))
                    unpaid_months.append((v, l, amount, 1.0))
                    month_amounts[v] = float(amount)
        except Teacher.DoesNotExist:
            pass

    salary_history = []
    if selected_teacher:
        salary_history = SalaryPayment.objects.filter(
            school=school, teacher=selected_teacher, academic_session=session,
        ).select_related('teacher').order_by('-payment_date')

    return render(request, 'school/salary/salary_form.html', {
        'form': form,
        'selected_teacher': selected_teacher,
        'unpaid_months': unpaid_months,
        'month_choices': available_month_choices,
        'month_amounts_json': json.dumps(month_amounts),
        'salary_history': salary_history,
        'session': session,
    })


@school_only
def salary_detail(request, pk):
    payment = get_object_or_404(SalaryPayment, pk=pk, school=request.user.school)
    return render(request, 'school/salary/salary_detail.html', {'payment': payment})


# ── Reports ───────────────────────────────────────────────────────────────────

def _style_header_cell(cell, bg_color='1E293B'):
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(fill_type='solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_data_cell(cell, row_idx):
    bg = 'F8FAFC' if row_idx % 2 == 0 else 'FFFFFF'
    cell.fill = PatternFill(fill_type='solid', fgColor=bg)
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    thin = Side(style='thin', color='E2E8F0')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _auto_width(ws, min_w=10, max_w=40, skip_rows=0):
    for col in ws.columns:
        cells = [cell for cell in col if cell.row > skip_rows]
        if not cells:
            continue
        length = max(len(str(cell.value or '')) for cell in cells)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


@school_only
def report_dashboard(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    classes = SchoolClass.objects.filter(school=school).order_by('name')
    session_choices = get_academic_session_choices(
        past_years=2, future_years=10,
        session_start_month=profile.session_start_month,
    )
    default_session = profile.current_academic_session
    return render(request, 'school/reports/report.html', {
        'profile': profile,
        'classes': classes,
        'session_choices': session_choices,
        'default_session': default_session,
    })


@school_only
def report_admissions_export(request):
    school = request.user.school
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str = request.GET.get('date_to', '').strip()
    class_id = request.GET.get('class_id', '').strip()
    session = request.GET.get('session', '').strip()

    qs = Student.objects.filter(school=school).select_related('school_class').order_by('admission_date', 'name')
    if session:
        qs = qs.filter(academic_session=session)

    if date_from_str:
        try:
            qs = qs.filter(admission_date__gte=date.fromisoformat(date_from_str))
        except ValueError:
            pass
    if date_to_str:
        try:
            qs = qs.filter(admission_date__lte=date.fromisoformat(date_to_str))
        except ValueError:
            pass
    if class_id:
        qs = qs.filter(school_class_id=class_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Admissions'
    ws.row_dimensions[1].height = 30

    # Column order matches the New Admission form sections exactly:
    # Student Details | Guardian Details | Additional Details | Transport
    headers = [
        # Student Details
        'Roll No.', 'Student Name', 'Date of Birth', 'Class', 'Academic Session',
        'Religion', 'Caste', 'Address', 'Admission Date',
        # Guardian Details
        'Father Name', 'Mother Name', 'WhatsApp Number',
        # Additional Details
        'Blood Group', 'Previous School', 'Aadhaar No.', 'PEN No.',
        # Transport
        'Transport Opted', 'Transport Amount (₹)',
        # Status
        'Status',
    ]
    ws.append(headers)
    for cell in ws[1]:
        _style_header_cell(cell)

    status_map = {'active': 'Active', 'inactive': 'Transferred', 'promoted': 'Promoted', 'fail': 'Fail'}

    for idx, s in enumerate(qs, start=2):
        ws.row_dimensions[idx].height = 18
        row = [
            # Student Details
            s.roll_number,
            s.name,
            s.date_of_birth.strftime('%d-%m-%Y') if s.date_of_birth else '',
            s.school_class.name,
            s.academic_session,
            s.religion or '',
            s.caste or '',
            s.address or '',
            s.admission_date.strftime('%d-%m-%Y') if s.admission_date else '',
            # Guardian Details
            s.father_name,
            s.mother_name,
            s.father_phone,
            # Additional Details
            s.blood_group or '',
            s.previous_school or '',
            s.aadhaar_number or '',
            s.pen_number or '',
            # Transport
            'Yes' if s.transport_opted else 'No',
            float(s.transport_amount) if s.transport_amount else '',
            # Status
            status_map.get(s.status, s.status),
        ]
        ws.append(row)
        for cell in ws[idx]:
            _style_data_cell(cell, idx)

    _auto_width(ws)
    ws.freeze_panes = 'A2'

    label_from = date_from_str or 'all'
    label_to = date_to_str or 'all'
    filename = f"admissions_{label_from}_to_{label_to}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@school_only
def report_fees_export(request):
    school = request.user.school
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str = request.GET.get('date_to', '').strip()
    class_id = request.GET.get('class_id', '').strip()
    session = request.GET.get('session', '').strip()

    qs = (
        FeePayment.objects
        .filter(school=school)
    )
    if session:
        qs = qs.filter(academic_session=session)
    qs = (
        qs
        .select_related('student', 'student__school_class', 'collected_by')
        .order_by('payment_date', 'student__name')
    )

    if date_from_str:
        try:
            qs = qs.filter(payment_date__gte=date.fromisoformat(date_from_str))
        except ValueError:
            pass
    if date_to_str:
        try:
            qs = qs.filter(payment_date__lte=date.fromisoformat(date_to_str))
        except ValueError:
            pass
    if class_id:
        qs = qs.filter(student__school_class_id=class_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fee Payments'
    ws.row_dimensions[1].height = 30

    headers = [
        'Receipt No.', 'Payment Date', 'Student Name', 'Roll No.', 'Class',
        'Academic Session', 'Months Paid', 'Amount Paid (₹)',
    ]
    ws.append(headers)
    for cell in ws[1]:
        _style_header_cell(cell, bg_color='0F4C81')

    month_map = dict(MONTH_CHOICES)

    for idx, p in enumerate(qs, start=2):
        ws.row_dimensions[idx].height = 18
        months_display = ', '.join(month_map.get(m, m) for m in (p.payment_months or []))
        row = [
            p.receipt_number,
            p.payment_date.strftime('%d-%m-%Y') if p.payment_date else '',
            p.student.name,
            p.student.roll_number,
            p.student.school_class.name,
            p.academic_session,
            months_display,
            float(p.amount_paid),
        ]
        ws.append(row)
        for cell in ws[idx]:
            _style_data_cell(cell, idx)

    # Totals row
    total_row_idx = qs.count() + 2
    ws.row_dimensions[total_row_idx].height = 20
    totals = qs.aggregate(
        paid=Sum('amount_paid'),
    )
    total_row = ['', '', '', '', '', '', 'TOTAL',
                 float(totals['paid'] or 0)]
    ws.append(total_row)
    for cell in ws[total_row_idx]:
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(fill_type='solid', fgColor='22C55E')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    _auto_width(ws)
    ws.freeze_panes = 'A2'

    label_from = date_from_str or 'all'
    label_to = date_to_str or 'all'
    filename = f"fees_{label_from}_to_{label_to}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _get_admission_qs(request):
    school = request.user.school
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str   = request.GET.get('date_to', '').strip()
    class_id      = request.GET.get('class_id', '').strip()
    session       = request.GET.get('session', '').strip()

    qs = Student.objects.filter(school=school).select_related('school_class').order_by('admission_date', 'name')
    if session:
        qs = qs.filter(academic_session=session)
    if date_from_str:
        try: qs = qs.filter(admission_date__gte=date.fromisoformat(date_from_str))
        except ValueError: pass
    if date_to_str:
        try: qs = qs.filter(admission_date__lte=date.fromisoformat(date_to_str))
        except ValueError: pass
    if class_id:
        qs = qs.filter(school_class_id=class_id)
    return qs, date_from_str, date_to_str, session


def _get_fees_qs(request):
    school = request.user.school
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str   = request.GET.get('date_to', '').strip()
    class_id      = request.GET.get('class_id', '').strip()
    session       = request.GET.get('session', '').strip()

    qs = FeePayment.objects.filter(school=school).select_related('student', 'student__school_class', 'collected_by')
    if session:
        qs = qs.filter(academic_session=session)
    if date_from_str:
        try: qs = qs.filter(payment_date__gte=date.fromisoformat(date_from_str))
        except ValueError: pass
    if date_to_str:
        try: qs = qs.filter(payment_date__lte=date.fromisoformat(date_to_str))
        except ValueError: pass
    if class_id:
        qs = qs.filter(student__school_class_id=class_id)
    qs = qs.order_by('payment_date', 'student__name')
    return qs, date_from_str, date_to_str, session


@school_only
def report_admissions_print(request):
    qs, date_from_str, date_to_str, session = _get_admission_qs(request)
    status_map = {'active': 'Active', 'inactive': 'Transferred', 'promoted': 'Promoted', 'fail': 'Fail'}
    students = [
        {
            'roll_number': s.roll_number,
            'name': s.name,
            'dob': s.date_of_birth.strftime('%d-%m-%Y') if s.date_of_birth else '',
            'class_name': s.school_class.name,
            'session': s.academic_session,
            'religion': s.religion or '',
            'caste': s.caste or '',
            'address': s.address or '',
            'admission_date': s.admission_date.strftime('%d-%m-%Y') if s.admission_date else '',
            'father_name': s.father_name,
            'mother_name': s.mother_name,
            'father_phone': s.father_phone,
            'blood_group': s.blood_group or '',
            'previous_school': s.previous_school or '',
            'aadhaar_number': s.aadhaar_number or '',
            'pen_number': s.pen_number or '',
            'transport': 'Yes' if s.transport_opted else 'No',
            'transport_amount': s.transport_amount or '',
            'status': status_map.get(s.status, s.status),
        }
        for s in qs
    ]
    return render(request, 'school/reports/print_admissions.html', {
        'students': students,
        'date_from': date_from_str,
        'date_to': date_to_str,
        'session': session,
        'school': request.user.school,
    })


@school_only
def report_fees_print(request):
    qs, date_from_str, date_to_str, session = _get_fees_qs(request)
    month_map = dict(MONTH_CHOICES)
    payments = [
        {
            'receipt_number': p.receipt_number,
            'payment_date': p.payment_date.strftime('%d-%m-%Y') if p.payment_date else '',
            'student_name': p.student.name,
            'roll_number': p.student.roll_number,
            'class_name': p.student.school_class.name,
            'session': p.academic_session,
            'months_paid': ', '.join(month_map.get(m, m) for m in (p.payment_months or [])),
            'gross_amount': p.gross_amount,
            'transport_amount': p.transport_amount,
            'amount_paid': p.amount_paid,
            'balance_due': p.balance_due,
            'collected_by': p.collected_by.username if p.collected_by else '',
        }
        for p in qs
    ]
    totals = qs.aggregate(
        gross=Sum('gross_amount'),
        transport=Sum('transport_amount'),
        paid=Sum('amount_paid'),
        balance=Sum('balance_due'),
    )
    return render(request, 'school/reports/print_fees.html', {
        'payments': payments,
        'totals': totals,
        'date_from': date_from_str,
        'date_to': date_to_str,
        'session': session,
        'school': request.user.school,
    })
