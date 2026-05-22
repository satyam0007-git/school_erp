from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def style_header_cell(cell, bg_color='1E293B'):
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(fill_type='solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_data_cell(cell, row_idx):
    bg = 'F8FAFC' if row_idx % 2 == 0 else 'FFFFFF'
    cell.fill = PatternFill(fill_type='solid', fgColor=bg)
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    thin = Side(style='thin', color='E2E8F0')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_column_widths(ws, min_w=10, max_w=40, skip_rows=0):
    for col in ws.columns:
        cells = [cell for cell in col if cell.row > skip_rows]
        if not cells:
            continue
        length = max(len(str(cell.value or '')) for cell in cells)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)
