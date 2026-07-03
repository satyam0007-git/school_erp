import io
import re
import logging

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from ..decorators import school_only
from ..forms import StudentForm
from ..logging_utils import log_activity_event
from ..models import AdmissionBulkUploadHistory, SchoolClass, SchoolProfile, Student, WhatsAppConfig

logger = logging.getLogger(__name__)
from ..services.student_service import (
    BULK_FIELD_INFO, BULK_UPLOAD_HEADERS,
    fail_student, process_bulk_upload, promote_student,
)
from .whatsapp_views import send_welcome_message_safely

from ..utils.excel_utils import set_column_widths, style_data_cell, style_header_cell


@school_only
def student_list(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    session = profile.current_academic_session

    q = request.GET.get('q', '').strip()
    class_id = request.GET.get('class')
    status = request.GET.get('status')
    transport = request.GET.get('transport')

    qs = Student.objects.filter(school=school, academic_session=session).select_related('school_class')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(roll_number__icontains=q) | Q(father_name__icontains=q))
    if class_id:
        qs = qs.filter(school_class_id=class_id)
    if status:
        qs = qs.filter(status=status)
    if transport == 'yes':
        qs = qs.filter(transport_opted=True)
    elif transport == 'no':
        qs = qs.filter(transport_opted=False)

    all_students = Student.objects.filter(school=school).select_related('school_class')
    session_class_ids = (
        Student.objects.filter(school=school, academic_session=session)
        .values_list('school_class_id', flat=True).distinct()
    )
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'school/admission/student_list.html', {
        'students': page_obj,
        'page_obj': page_obj,
        'is_paginated': paginator.num_pages > 1,
        'classes': SchoolClass.objects.filter(pk__in=session_class_ids).order_by('name'),
        'status_choices': Student.STATUS_CHOICES,
        'session_start_month': profile.session_start_month.capitalize(),
        'session_end_month': profile.session_end_month.capitalize(),
        'total_students': all_students.count(),
        'session_wise': all_students.values('academic_session').annotate(count=Count('id')).order_by('-academic_session'),
        'filtered_total': qs.count(),
        'filtered_active': qs.filter(status='active').count(),
        'filtered_promoted': qs.filter(status='promoted').count(),
        'filtered_inactive': qs.filter(status='inactive').count(),
        'filtered_fail': qs.filter(status='fail').count(),
        'selected_class': class_id or '',
        'selected_status': status or '',
        'selected_transport': transport or '',
        'selected_q': q,
        'has_filters': bool(q or class_id or status or transport in ['yes', 'no']),
    })


@school_only
def student_create(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    session = profile.current_academic_session
    form = StudentForm(request.POST or None, school=school, session=session)
    if request.method == 'POST' and form.is_valid():
        student = form.save(commit=False)
        student.school = school
        student.academic_session = session
        student.save()
        discount_months = form.cleaned_data.get('discount_months') or 0
        log_activity_event(
            request,
            module='student',
            action='create',
            record_id=student.pk,
            details={
                'name': student.name,
                'class': student.school_class_id,
                'academic_session': student.academic_session,
                'discount_months': discount_months,
            },
        )

        # Trigger WhatsApp welcome message if option is selected
        send_welcome = form.cleaned_data.get('send_whatsapp_welcome', False)
        if send_welcome:
            send_welcome_message_safely(school, student, request=request)

        msg = f'Student admitted. First {discount_months} month(s) discounted.' if discount_months > 0 else 'Student admitted.'
        messages.success(request, msg)
        return redirect('student_list')
    return render(request, 'school/admission/student_form.html', {'form': form})


@school_only
def student_edit(request, pk):
    school = request.user.school
    student = get_object_or_404(Student, pk=pk, school=school)
    form = StudentForm(request.POST or None, instance=student, school=school, session=student.academic_session)
    if request.method == 'POST' and form.is_valid():
        old_values = {
            'name': student.name,
            'school_class_id': student.school_class_id,
            'father_name': student.father_name,
            'mother_name': student.mother_name,
            'father_phone': student.father_phone,
            'address': student.address,
            'religion': student.religion,
            'caste': student.caste,
            'blood_group': student.blood_group,
            'transport_opted': student.transport_opted,
            'transport_amount': str(student.transport_amount) if student.transport_amount is not None else None,
            'discount_months': student.discount_months,
        }
        form.save()
        log_activity_event(
            request,
            module='student',
            action='update',
            record_id=student.pk,
            old_values=old_values,
            new_values={
                'name': student.name,
                'school_class_id': student.school_class_id,
                'father_name': student.father_name,
                'mother_name': student.mother_name,
                'father_phone': student.father_phone,
                'address': student.address,
                'religion': student.religion,
                'caste': student.caste,
                'blood_group': student.blood_group,
                'transport_opted': student.transport_opted,
                'transport_amount': str(student.transport_amount) if student.transport_amount is not None else None,
                'discount_months': student.discount_months,
            },
        )
        messages.success(request, 'Student updated.')
        return redirect('student_list')
    return render(request, 'school/admission/student_form.html', {'form': form, 'object': student})


@school_only
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk, school=request.user.school)
    if request.method == 'POST':
        try:
            student_snapshot = {
                'name': student.name,
                'roll_number': student.roll_number,
                'class': student.school_class_id,
            }
            student.delete()
            log_activity_event(
                request,
                module='student',
                action='delete',
                record_id=pk,
                details=student_snapshot,
            )
            messages.success(request, 'Student deleted.')
        except ProtectedError:
            payment_count = student.fee_payments.count()
            log_activity_event(
                request,
                module='student',
                action='delete',
                record_id=pk,
                status='failure',
                details={'name': student.name, 'reason': 'protected_by_fee_payments', 'payment_count': payment_count},
            )
            messages.error(
                request,
                f'Cannot delete {student.name}: they have {payment_count} fee payment record(s). '
                'Delete the payments first.',
            )
        return redirect('student_list')
    return render(request, 'school/admission/confirm_delete.html', {'object': student})


@school_only
def student_promote(request, pk):
    if request.method != 'POST':
        return redirect('student_list')
    school = request.user.school
    student = get_object_or_404(Student, pk=pk, school=school, status=Student.STATUS_ACTIVE)
    promoted, error = promote_student(student, school)
    if error:
        log_activity_event(
            request,
            module='student',
            action='promote',
            record_id=student.pk,
            status='failure',
            details={'name': student.name, 'reason': error},
        )
        messages.error(request, error)
    else:
        log_activity_event(
            request,
            module='student',
            action='promote',
            record_id=student.pk,
            details={
                'name': student.name,
                'promoted_to_student_id': promoted.pk,
                'new_class': promoted.school_class_id,
                'new_session': promoted.academic_session,
            },
        )
        messages.success(
            request,
            f'{student.name} promoted → {promoted.school_class.name} ({promoted.academic_session}).',
        )
    return redirect('student_list')


@school_only
def student_fail(request, pk):
    if request.method != 'POST':
        return redirect('student_list')
    school = request.user.school
    student = get_object_or_404(Student, pk=pk, school=school, status=Student.STATUS_ACTIVE)
    retained = fail_student(student)
    log_activity_event(
        request,
        module='student',
        action='fail',
        record_id=student.pk,
        details={'name': student.name, 'retained_student_id': retained.pk, 'class': retained.school_class_id, 'session': retained.academic_session},
    )
    messages.success(
        request,
        f'{student.name} marked as failed and retained in {retained.school_class.name} ({retained.academic_session}).',
    )
    return redirect('student_list')


@school_only
def student_transfer(request, pk):
    if request.method != 'POST':
        return redirect('student_list')
    student = get_object_or_404(Student, pk=pk, school=request.user.school, status=Student.STATUS_ACTIVE)
    previous_status = student.status
    student.status = Student.STATUS_INACTIVE
    student.save(update_fields=['status', 'updated_at'])
    log_activity_event(
        request,
        module='student',
        action='transfer',
        record_id=student.pk,
        old_values={'status': previous_status},
        new_values={'status': student.status},
        details={'name': student.name},
    )
    messages.success(request, f'{student.name} marked as transferred.')
    return redirect('student_list')


@school_only
def admission_bulk_template(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    session = profile.current_academic_session
    
    # Get classes that have students in the current session
    classes_with_students = list(
        SchoolClass.objects
        .filter(students__school=school, students__academic_session=session)
        .distinct()
        .order_by('name')
    )
    
    # If no students in session, get all classes for school
    if not classes_with_students:
        classes_with_students = list(SchoolClass.objects.filter(school=school).order_by('name'))
    
    classes = classes_with_students

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bulk Admission'
    ws.row_dimensions[1].height = 38
    ws.append(BULK_UPLOAD_HEADERS)
    for cell in ws[1]:
        style_header_cell(cell)

    ws_lists = wb.create_sheet('Lists')
    ws_lists.sheet_state = 'hidden'
    for i, cls in enumerate(classes, start=1):
        ws_lists.cell(row=i, column=1, value=cls.name)

    def add_dropdown(formula1, col_letter):
        dv = DataValidation(type='list', formula1=formula1, allow_blank=True, showDropDown=False)
        dv.sqref = f'{col_letter}2:{col_letter}201'
        ws.add_data_validation(dv)

    if classes:
        add_dropdown(f'Lists!$A$1:$A${len(classes)}', 'C')
    add_dropdown('"Hindu,Muslim,Christian,Sikh,Buddhist,Jain,Parsi,Other"', 'H')
    add_dropdown('"General,OBC,SC,ST,EWS,Other"', 'I')
    add_dropdown('"O+,O-,A+,A-,B+,B-,AB+,AB-"', 'J')
    add_dropdown('"Yes,No"', 'N')
    add_dropdown('"Yes,No"', 'T')

    for col_idx in [2, 17, 19]:
        for row_num in range(1, 203):
            ws.cell(row=row_num, column=col_idx).number_format = '@'

    set_column_widths(ws, max_w=30)
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bulk_admission_template.xlsx"'
    return response


@school_only
def admission_bulk_upload(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    classes = SchoolClass.objects.filter(school=school).order_by('name')
    upload_result = None

    if request.method == 'GET':
        request.session.pop('bulk_upload_failed', None)

    if request.method == 'POST':
        uploaded_file = request.FILES.get('excel_file')
        if not uploaded_file:
            messages.error(request, 'Please select an Excel file to upload.')
        elif not uploaded_file.name.lower().endswith('.xlsx'):
            messages.error(request, 'Only .xlsx files are supported.')
        else:
            try:
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                upload_result, failed_rows = process_bulk_upload(wb, school, profile, request.user)
                history = AdmissionBulkUploadHistory.objects.create(
                    school=school,
                    uploaded_by=request.user,
                    academic_session=profile.current_academic_session,
                    file_name=uploaded_file.name[:255],
                    total_records=upload_result['total'],
                    admissions_created=upload_result['success'],
                    fee_submissions=upload_result['fee_success'],
                    failed_records=upload_result['failed'],
                    fee_skipped=upload_result['fee_skipped'],
                )
                Student.objects.filter(
                    school=school,
                    pk__in=upload_result.get('successful_student_ids', []),
                ).update(bulk_upload_history=history)
                if failed_rows:
                    request.session['bulk_upload_failed'] = failed_rows
                else:
                    request.session.pop('bulk_upload_failed', None)
                log_activity_event(
                    request,
                    module='student',
                    action='bulk_upload',
                    details={
                        'file_name': uploaded_file.name,
                        'total_records': upload_result['total'],
                        'success': upload_result['success'],
                        'failed': upload_result['failed'],
                        'fee_success': upload_result['fee_success'],
                        'fee_skipped': upload_result['fee_skipped'],
                    },
                )
            except Exception as exc:
                log_activity_event(
                    request,
                    module='student',
                    action='bulk_upload',
                    status='failure',
                    details={'file_name': uploaded_file.name if uploaded_file else '', 'error': str(exc)},
                )
                messages.error(request, f'Failed to process file: {exc}')

    upload_history = AdmissionBulkUploadHistory.objects.filter(
        school=school,
        academic_session=profile.current_academic_session,
    ).select_related('uploaded_by')[:20]

    return render(request, 'school/admission/bulk_upload.html', {
        'classes': classes,
        'field_info': BULK_FIELD_INFO,
        'upload_result': upload_result,
        'upload_history': upload_history,
        'has_pending_errors': bool(request.session.get('bulk_upload_failed')),
        'current_session': profile.current_academic_session,
    })


@school_only
def admission_bulk_success_download(request, pk):
    school = request.user.school
    history = get_object_or_404(AdmissionBulkUploadHistory, pk=pk, school=school)
    students = (
        Student.objects
        .filter(school=school, bulk_upload_history=history)
        .select_related('school_class')
        .order_by('school_class__name', 'name')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Successful Admissions'
    ws.row_dimensions[1].height = 30

    headers = [
        'Roll No.', 'Student Name', 'Date of Birth', 'Class', 'Academic Session',
        'Religion', 'Caste', 'Address', 'Admission Date',
        'Father Name', 'Mother Name', 'WhatsApp Number',
        'Blood Group', 'Previous School', 'Aadhaar No.', 'PEN No.',
        'Transport Opted', 'Transport Amount',
    ]
    ws.append(headers)
    for cell in ws[1]:
        style_header_cell(cell, '15803D')

    for row_idx, student in enumerate(students, start=2):
        ws.row_dimensions[row_idx].height = 18
        ws.append([
            student.roll_number,
            student.name,
            student.date_of_birth.strftime('%d-%m-%Y') if student.date_of_birth else '',
            student.school_class.name,
            student.academic_session,
            student.religion or '',
            student.caste or '',
            student.address or '',
            student.admission_date.strftime('%d-%m-%Y') if student.admission_date else '',
            student.father_name,
            student.mother_name,
            student.father_phone,
            student.blood_group or '',
            student.previous_school or '',
            student.aadhaar_number or '',
            student.pen_number or '',
            'Yes' if student.transport_opted else 'No',
            float(student.transport_amount) if student.transport_amount else '',
        ])
        for cell in ws[row_idx]:
            style_data_cell(cell, row_idx)

    set_column_widths(ws)
    ws.freeze_panes = 'A2'

    safe_file = re.sub(r'[^\w.-]+', '_', history.file_name or 'bulk_upload').strip('._')
    filename = f"successful_admissions_{history.pk}_{safe_file or 'bulk_upload'}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@school_only
def admission_bulk_errors_download(request):
    failed_rows = request.session.get('bulk_upload_failed', [])
    headers = [
        'Student Name', 'Date of Birth', 'Class', 'Father Name', 'Mother Name',
        'Father WhatsApp', 'Address', 'Religion', 'Caste', 'Blood Group',
        'Previous School', 'Aadhaar Number', 'PEN Number', 'Transport Opted',
        'Transport Amount', 'Discount Months', 'Admission Date', 'Paid Amount',
        'Payment Date', 'Send WhatsApp Welcome', 'Error Remarks',
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Failed Records'
    ws.row_dimensions[1].height = 30
    ws.append(headers)
    thin = Side(style='thin', color='E2E8F0')
    for col_idx, cell in enumerate(ws[1], start=1):
        style_header_cell(cell, bg_color='DC2626' if col_idx == len(headers) else '1E293B')

    for col_idx in [2, 17, 19]:
        for row_num in range(1, len(failed_rows) + 10):
            ws.cell(row=row_num, column=col_idx).number_format = '@'

    for row_idx, row_data in enumerate(failed_rows, start=2):
        ws.row_dimensions[row_idx].height = 22
        padded = list(row_data) + [''] * max(0, len(headers) - len(row_data))
        ws.append(padded[:len(headers)])
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            if col_idx == len(headers):
                cell.fill = PatternFill(fill_type='solid', fgColor='FEE2E2')
                cell.font = Font(color='DC2626', size=10)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                style_data_cell(cell, row_idx)
            if col_idx in (2, 18, 21):
                cell.number_format = '@'

    set_column_widths(ws, max_w=45)
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bulk_upload_errors.xlsx"'
    return response
