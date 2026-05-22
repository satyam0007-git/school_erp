import io
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from ..decorators import school_only
from ..models import FeePayment, SchoolClass, SchoolProfile, Student, MONTH_CHOICES
from ..utils.excel_utils import set_column_widths, style_data_cell, style_header_cell


@school_only
def report_dashboard(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    classes = SchoolClass.objects.filter(school=school).order_by('name')
    return render(request, 'school/reports/report.html', {
        'profile': profile,
        'classes': classes,
    })


@school_only
def report_admissions_export(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str = request.GET.get('date_to', '').strip()
    class_id = request.GET.get('class_id', '').strip()
    session = profile.current_academic_session

    qs = Student.objects.filter(
        school=school, academic_session=session,
    ).select_related('school_class').order_by('admission_date', 'name')
    if date_from_str:
        try:
            qs = qs.filter(admission_date__gte=date.fromisoformat(date_from_str))
        except ValueError:
            pass
    if date_to_str:
        try:
            qs = qs.filter(admission_date__lte=date.fromisoformat(date_to_str))
        except ValueError:
            pass
    if class_id:
        qs = qs.filter(school_class_id=class_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Admissions'
    ws.row_dimensions[1].height = 30

    headers = [
        'Roll No.', 'Student Name', 'Date of Birth', 'Class', 'Academic Session',
        'Religion', 'Caste', 'Address', 'Admission Date',
        'Father Name', 'Mother Name', 'WhatsApp Number',
        'Blood Group', 'Previous School', 'Aadhaar No.', 'PEN No.',
        'Transport Opted', 'Transport Amount (₹)',
        'Status',
    ]
    ws.append(headers)
    for cell in ws[1]:
        style_header_cell(cell)

    status_map = {'active': 'Active', 'inactive': 'Transferred', 'promoted': 'Promoted', 'fail': 'Fail'}
    for idx, s in enumerate(qs, start=2):
        ws.row_dimensions[idx].height = 18
        ws.append([
            s.roll_number, s.name,
            s.date_of_birth.strftime('%d-%m-%Y') if s.date_of_birth else '',
            s.school_class.name, s.academic_session,
            s.religion or '', s.caste or '', s.address or '',
            s.admission_date.strftime('%d-%m-%Y') if s.admission_date else '',
            s.father_name, s.mother_name, s.father_phone,
            s.blood_group or '', s.previous_school or '', s.aadhaar_number or '', s.pen_number or '',
            'Yes' if s.transport_opted else 'No',
            float(s.transport_amount) if s.transport_amount else '',
            status_map.get(s.status, s.status),
        ])
        for cell in ws[idx]:
            style_data_cell(cell, idx)

    set_column_widths(ws)
    ws.freeze_panes = 'A2'

    filename = f"admissions_{date_from_str or 'all'}_to_{date_to_str or 'all'}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@school_only
def report_fees_export(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str = request.GET.get('date_to', '').strip()
    class_id = request.GET.get('class_id', '').strip()
    session = profile.current_academic_session

    qs = (
        FeePayment.objects
        .filter(school=school, academic_session=session)
        .select_related('student', 'student__school_class', 'collected_by')
        .order_by('payment_date', 'student__name')
    )
    if date_from_str:
        try:
            qs = qs.filter(payment_date__gte=date.fromisoformat(date_from_str))
        except ValueError:
            pass
    if date_to_str:
        try:
            qs = qs.filter(payment_date__lte=date.fromisoformat(date_to_str))
        except ValueError:
            pass
    if class_id:
        qs = qs.filter(student__school_class_id=class_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fee Payments'
    ws.row_dimensions[1].height = 30

    headers = [
        'Receipt No.', 'Payment Date', 'Student Name', 'Roll No.', 'Class',
        'Academic Session', 'Months Paid', 'Amount Paid (₹)',
    ]
    ws.append(headers)
    for cell in ws[1]:
        style_header_cell(cell, '0F4C81')

    month_map = dict(MONTH_CHOICES)
    for idx, p in enumerate(qs, start=2):
        ws.row_dimensions[idx].height = 18
        months_display = ', '.join(month_map.get(m, m) for m in (p.payment_months or []))
        ws.append([
            p.receipt_number,
            p.payment_date.strftime('%d-%m-%Y') if p.payment_date else '',
            p.student.name, p.student.roll_number, p.student.school_class.name,
            p.academic_session, months_display, float(p.amount_paid),
        ])
        for cell in ws[idx]:
            style_data_cell(cell, idx)

    total_row_idx = qs.count() + 2
    ws.row_dimensions[total_row_idx].height = 20
    totals = qs.aggregate(paid=Sum('amount_paid'))
    ws.append(['', '', '', '', '', '', 'TOTAL', float(totals['paid'] or 0)])
    for cell in ws[total_row_idx]:
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(fill_type='solid', fgColor='22C55E')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    set_column_widths(ws)
    ws.freeze_panes = 'A2'

    filename = f"fees_{date_from_str or 'all'}_to_{date_to_str or 'all'}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@school_only
def report_admissions_print(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str = request.GET.get('date_to', '').strip()
    class_id = request.GET.get('class_id', '').strip()
    session = profile.current_academic_session

    qs = Student.objects.filter(
        school=school, academic_session=session,
    ).select_related('school_class').order_by('admission_date', 'name')
    if date_from_str:
        try:
            qs = qs.filter(admission_date__gte=date.fromisoformat(date_from_str))
        except ValueError:
            pass
    if date_to_str:
        try:
            qs = qs.filter(admission_date__lte=date.fromisoformat(date_to_str))
        except ValueError:
            pass
    if class_id:
        qs = qs.filter(school_class_id=class_id)

    status_map = {'active': 'Active', 'inactive': 'Transferred', 'promoted': 'Promoted', 'fail': 'Fail'}
    students = [
        {
            'roll_number': s.roll_number, 'name': s.name,
            'dob': s.date_of_birth.strftime('%d-%m-%Y') if s.date_of_birth else '',
            'class_name': s.school_class.name, 'session': s.academic_session,
            'religion': s.religion or '', 'caste': s.caste or '', 'address': s.address or '',
            'admission_date': s.admission_date.strftime('%d-%m-%Y') if s.admission_date else '',
            'father_name': s.father_name, 'mother_name': s.mother_name, 'father_phone': s.father_phone,
            'blood_group': s.blood_group or '', 'previous_school': s.previous_school or '',
            'aadhaar_number': s.aadhaar_number or '', 'pen_number': s.pen_number or '',
            'transport': 'Yes' if s.transport_opted else 'No',
            'transport_amount': s.transport_amount or '',
            'status': status_map.get(s.status, s.status),
        }
        for s in qs
    ]
    return render(request, 'school/reports/print_admissions.html', {
        'students': students,
        'date_from': date_from_str,
        'date_to': date_to_str,
        'session': session,
        'school': school,
    })


@school_only
def report_fees_print(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str = request.GET.get('date_to', '').strip()
    class_id = request.GET.get('class_id', '').strip()
    session = profile.current_academic_session

    qs = (
        FeePayment.objects.filter(school=school, academic_session=session)
        .select_related('student', 'student__school_class', 'collected_by')
    )
    if date_from_str:
        try:
            qs = qs.filter(payment_date__gte=date.fromisoformat(date_from_str))
        except ValueError:
            pass
    if date_to_str:
        try:
            qs = qs.filter(payment_date__lte=date.fromisoformat(date_to_str))
        except ValueError:
            pass
    if class_id:
        qs = qs.filter(student__school_class_id=class_id)
    qs = qs.order_by('payment_date', 'student__name')

    month_map = dict(MONTH_CHOICES)
    payments = [
        {
            'receipt_number': p.receipt_number,
            'payment_date': p.payment_date.strftime('%d-%m-%Y') if p.payment_date else '',
            'student_name': p.student.name,
            'roll_number': p.student.roll_number,
            'class_name': p.student.school_class.name,
            'session': p.academic_session,
            'months_paid': ', '.join(month_map.get(m, m) for m in (p.payment_months or [])),
            'gross_amount': p.gross_amount,
            'transport_amount': p.transport_amount,
            'amount_paid': p.amount_paid,
            'balance_due': p.balance_due,
            'collected_by': p.collected_by.username if p.collected_by else '',
        }
        for p in qs
    ]
    totals = qs.aggregate(
        gross=Sum('gross_amount'),
        transport=Sum('transport_amount'),
        paid=Sum('amount_paid'),
        balance=Sum('balance_due'),
    )
    return render(request, 'school/reports/print_fees.html', {
        'payments': payments,
        'totals': totals,
        'date_from': date_from_str,
        'date_to': date_to_str,
        'session': session,
        'school': school,
    })
