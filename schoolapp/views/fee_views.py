import calendar
import io
import re
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Max, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from ..decorators import school_only
from ..forms import FeePaymentForm
from ..models import (
    ExamFee, FeePayment, FeeStructure, SchoolClass, SchoolProfile,
    SchoolSessionRecord, Student, WhatsAppConfig, MONTH_CHOICES,
)
from ..services.fee_service import (
    distribute_lump_sum, get_available_advance, get_editable_fee_options,
    get_monthly_tuition_fee, get_transport_fee, get_unpaid_fee_options,
    get_unpaid_exam_fees,
)
from ..services.report_service import build_fee_dashboard_data
from ..session_utils import CAL_TO_MONTH, MONTH_TO_CAL, get_session_months
from ..utils.excel_utils import set_column_widths, style_data_cell, style_header_cell


@school_only
def payment_dashboard(request):
    school = request.user.school
    reference_date = timezone.localdate()
    profile = SchoolProfile.get_for_school(school)
    selected_session = profile.current_academic_session

    q = request.GET.get('q', '').strip()
    class_id = request.GET.get('class')
    payment_status = request.GET.get('payment_status', '')

    session_record = SchoolSessionRecord.objects.filter(
        school=school, academic_session=selected_session
    ).first()
    session_start_month = session_record.session_start_month if session_record else profile.session_start_month
    session_end_month = session_record.session_end_month if session_record else profile.session_end_month

    students = Student.objects.filter(
        school=school, status=Student.STATUS_ACTIVE, academic_session=selected_session,
    ).select_related('school_class')
    if q:
        students = students.filter(
            Q(name__icontains=q) | Q(father_name__icontains=q) | Q(roll_number__icontains=q)
        )
    if class_id:
        students = students.filter(school_class_id=class_id)

    session_months = get_session_months(session_start_month, session_end_month)
    current_month_key = CAL_TO_MONTH[reference_date.month]

    try:
        sess_start_year = int(selected_session.split('-')[0])
    except (ValueError, IndexError):
        sess_start_year = reference_date.year
    sess_start_cal = MONTH_TO_CAL.get(session_start_month, 1)
    sess_end_cal = MONTH_TO_CAL.get(session_end_month, 3)
    sess_end_year = sess_start_year if sess_end_cal >= sess_start_cal else sess_start_year + 1
    sess_start_date = date(sess_start_year, sess_start_cal, 1)
    sess_end_date = date(sess_end_year, sess_end_cal, calendar.monthrange(sess_end_year, sess_end_cal)[1])

    if reference_date < sess_start_date:
        months_due = 0
    elif reference_date > sess_end_date:
        months_due = len(session_months)
    elif current_month_key in session_months:
        months_due = session_months.index(current_month_key) + 1
    else:
        months_due = len(session_months)

    month_label_map = {v: l for v, l in MONTH_CHOICES}

    latest_payment_by_student = dict(
        FeePayment.objects.filter(school=school, academic_session=selected_session)
        .values('student_id')
        .annotate(latest_id=Max('id'))
        .values_list('student_id', 'latest_id')
    )

    session_payments = FeePayment.objects.filter(
        school=school, academic_session=selected_session,
    ).values('student_id', 'payment_months', 'exam_fee_items', 'payment_date')
    paid_months_by_student = {}
    exam_items_by_student = {}
    for p in session_payments:
        sid = p['student_id']
        paid_months_by_student.setdefault(sid, [])
        exam_items_by_student.setdefault(sid, [])
        date_str = p['payment_date'].strftime('%d %b %Y') if p['payment_date'] else '—'
        for m in (p['payment_months'] or []):
            m = str(m)
            if m.endswith('_transport'):
                base = m.replace('_transport', '')
                label = month_label_map.get(base, base)
                paid_months_by_student[sid].append({'label': label, 'transport': True, 'date': date_str})
            else:
                label = month_label_map.get(m, m)
                paid_months_by_student[sid].append({'label': label, 'transport': False, 'date': date_str})
        for item in (p['exam_fee_items'] or []):
            name = item.get('name') or item.get('exam_name') or str(item)
            entry = {'name': name, 'date': date_str}
            if not any(e['name'] == name for e in exam_items_by_student[sid]):
                exam_items_by_student[sid].append(entry)

    payment_data = []
    for student in students.order_by('school_class__name', 'name'):
        monthly_fee = get_monthly_tuition_fee(student, school, session=selected_session)
        transport_fee = get_transport_fee(student)
        exam_fees = get_unpaid_exam_fees(student, school, session=selected_session)
        total_needed = (monthly_fee + transport_fee) * months_due + exam_fees
        total_paid = FeePayment.objects.filter(
            student=student, academic_session=selected_session,
        ).aggregate(t=Sum('amount_paid'))['t'] or Decimal('0.00')
        balance = max(total_needed - total_paid, Decimal('0.00'))
        payment_data.append({
            'student_id': student.id,
            'student_name': student.name,
            'roll_number': student.roll_number,
            'father_name': student.father_name,
            'student_class': str(student.school_class),
            'father_phone': student.father_phone,
            'total_amount_paid': total_paid,
            'total_payment_need_to_pay_till_month': total_needed,
            'balance_payment': balance,
            'is_paid_up': balance <= Decimal('0.00'),
            'paid_months': paid_months_by_student.get(student.id, []),
            'exam_items': exam_items_by_student.get(student.id, []),
            'latest_payment_id': latest_payment_by_student.get(student.id),
        })

    if payment_status == 'pending':
        payment_data = [p for p in payment_data if p['balance_payment'] > 0]
    elif payment_status == 'paid':
        payment_data = [p for p in payment_data if p['balance_payment'] <= 0]

    total_due = sum(p['total_payment_need_to_pay_till_month'] for p in payment_data)
    total_collected = sum(p['total_amount_paid'] for p in payment_data)
    total_pending = sum(p['balance_payment'] for p in payment_data if p['balance_payment'] > 0)

    paginator = Paginator(payment_data, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    if months_due == 0:
        fee_range_label = '—'
    else:
        try:
            label_start_year = int(selected_session.split('-')[0])
        except (ValueError, IndexError):
            label_start_year = reference_date.year
        label_start_cal = MONTH_TO_CAL.get(session_start_month, 1)
        range_start_key = session_months[0] if session_months else session_start_month
        range_end_key = session_months[months_due - 1]
        year_for = lambda key: label_start_year if MONTH_TO_CAL[key] >= label_start_cal else label_start_year + 1
        fee_range_label = (
            f"{range_start_key.capitalize()} {year_for(range_start_key)} → "
            f"{range_end_key.capitalize()} {year_for(range_end_key)}"
        )

    wa_config, _ = WhatsAppConfig.objects.get_or_create(school=school)

    session_class_ids = FeeStructure.objects.filter(
        fee_category__school=school,
        academic_session=selected_session,
        frequency=FeeStructure.FREQUENCY_MONTHLY,
    ).values_list('school_class_id', flat=True)
    session_classes = SchoolClass.objects.filter(school=school, pk__in=session_class_ids).order_by('name')

    return render(request, 'school/fees/payment_dashboard.html', {
        'payment_summary': page_obj,
        'page_obj': page_obj,
        'is_paginated': paginator.num_pages > 1,
        'classes': session_classes,
        'selected_session': selected_session,
        'selected_class': class_id or '',
        'selected_q': q,
        'selected_payment_status': payment_status,
        'has_filters': bool(q or class_id or payment_status),
        'total_students': len(payment_data),
        'total_due': total_due,
        'total_collected': total_collected,
        'total_pending': total_pending,
        'wa_enabled': wa_config.is_active,
        'fee_range_label': fee_range_label,
    })


@school_only
def payment_dashboard_export(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    filters = {
        'q': request.GET.get('q', '').strip(),
        'class_id': request.GET.get('class'),
        'payment_status': request.GET.get('payment_status', ''),
        'session': request.GET.get('session', '').strip(),
    }
    data = build_fee_dashboard_data(school, profile, filters)
    payment_data = data['payment_data']
    selected_session = data['selected_session']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fee Submission'

    num_cols = 8
    last_col = get_column_letter(num_cols)
    for merged in (f'A1:{last_col}1', f'A2:{last_col}2', f'A3:{last_col}3'):
        ws.merge_cells(merged)
    ws['A1'].value = f"School: {school.name}"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A2'].value = f"Session: {selected_session}"
    ws['A2'].font = Font(size=10, color='64748B')
    ws['A3'].value = f"Billing Period: {data['fee_range_label']}"
    ws['A3'].font = Font(size=10, color='0D9488')
    for r, h in [(1, 18), (2, 16), (3, 16), (4, 6), (5, 30)]:
        ws.row_dimensions[r].height = h

    headers = ['S.N', 'Student Name', "Father's Name", 'Class', 'Total Due (₹)', 'Paid (₹)', 'Balance (₹)', 'Status']
    for col, h in enumerate(headers, 1):
        style_header_cell(ws.cell(row=5, column=col, value=h), '1E40AF')

    for idx, p in enumerate(payment_data, 6):
        ws.append([
            idx - 5, p['student_name'], p['father_name'], p['student_class'],
            float(p['total_payment_need_to_pay_till_month']),
            float(p['total_amount_paid']),
            float(p['balance_payment']),
            'Paid' if p['is_paid_up'] else 'Pending',
        ])
        for cell in ws[idx]:
            style_data_cell(cell, idx)

    summary_row = ws.max_row + 1
    ws.cell(row=summary_row, column=1, value='TOTAL')
    ws.cell(row=summary_row, column=5, value=float(data['total_due']))
    ws.cell(row=summary_row, column=6, value=float(data['total_collected']))
    ws.cell(row=summary_row, column=7, value=float(data['total_pending']))
    thin = Side(style='thin', color='BFDBFE')
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=summary_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='EFF6FF')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    set_column_widths(ws, skip_rows=4)
    ws.freeze_panes = 'A6'

    safe_name = re.sub(r'[^\w\-]', '_', school.name)
    filename = f"fee_submission_{safe_name}_{selected_session}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@school_only
def payment_dashboard_print(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    filters = {
        'q': request.GET.get('q', '').strip(),
        'class_id': request.GET.get('class'),
        'payment_status': request.GET.get('payment_status', ''),
        'session': request.GET.get('session', '').strip(),
    }
    data = build_fee_dashboard_data(school, profile, filters)
    return render(request, 'school/fees/print_fee_dashboard.html', {
        'payment_data': data['payment_data'],
        'total_due': data['total_due'],
        'total_collected': data['total_collected'],
        'total_pending': data['total_pending'],
        'selected_session': data['selected_session'],
        'school': data['school'],
        'total_students': len(data['payment_data']),
        'printed_on': timezone.localdate(),
        'fee_range_label': data['fee_range_label'],
    })


@school_only
def payment_create(request):
    school = request.user.school
    profile = SchoolProfile.get_for_school(school)
    current_session = profile.current_academic_session

    student_qs = Student.objects.filter(
        school=school, status=Student.STATUS_ACTIVE, academic_session=current_session,
    ).select_related('school_class')

    q = request.GET.get('q', '').strip()
    class_id = request.GET.get('class')
    filtered_qs = student_qs
    if q:
        filtered_qs = filtered_qs.filter(Q(name__icontains=q) | Q(roll_number__icontains=q) | Q(father_name__icontains=q))
    if class_id:
        filtered_qs = filtered_qs.filter(school_class_id=class_id)

    has_active_filters = bool(q or class_id)
    paginator = Paginator(filtered_qs.order_by('name'), 10)
    filtered_page = paginator.get_page(request.GET.get('page'))

    if request.method == 'POST':
        payment_mode = request.POST.get('payment_mode', 'manual')
        form = FeePaymentForm(request.POST, student_queryset=student_qs, lump_sum_mode=(payment_mode == 'lump_sum'))

        if form.is_valid():
            with transaction.atomic():
                payment = form.save(commit=False)
                payment.school = school
                payment.academic_session = payment.student.academic_session
                payment.collected_by = request.user

                if payment_mode == 'lump_sum':
                    try:
                        cash_amount = Decimal(str(request.POST.get('lump_sum_amount', '0')))
                    except InvalidOperation:
                        cash_amount = Decimal('0.00')

                    if cash_amount <= 0:
                        form.add_error(None, 'Enter a valid lump-sum amount greater than zero.')
                    else:
                        advance_available = get_available_advance(payment.student, payment.student.academic_session)
                        result = distribute_lump_sum(payment.student, school, cash_amount, advance_available, profile)
                        if not result['paid_month_tokens'] and not result['exam_fee_items']:
                            form.add_error(None, 'All fees are already paid for this student or no fee structure is set up.')
                        else:
                            payment.is_lump_sum = True
                            payment.payment_months = result['paid_month_tokens']
                            payment.exam_fee_items = result['exam_fee_items']
                            payment.transport_amount = result['transport_total']
                            payment.amount_paid = cash_amount
                            payment.gross_amount = result['gross_amount']
                            payment.advance_used = advance_available
                            payment.advance_balance = result['advance_balance']
                            payment.save()
                            messages.success(request, 'Lump-sum payment saved and auto-distributed.')
                            return redirect('payment_dashboard')
                else:
                    selected_items = form.cleaned_data['payment_months']
                    selected_months = [m for m in selected_items if not str(m).startswith('exam_')]
                    selected_exam_keys = [m for m in selected_items if str(m).startswith('exam_')]

                    unpaid_options = get_unpaid_fee_options(payment.student, school)
                    unpaid_values = {item['value'] for item in unpaid_options}
                    invalid = [m for m in selected_items if m not in unpaid_values]
                    if invalid:
                        form.add_error('payment_months', 'Some selected items are already paid.')
                    else:
                        transport_count = sum(1 for m in selected_months if str(m).endswith('_transport'))
                        base_count = len(selected_months) - transport_count
                        monthly_base = get_monthly_tuition_fee(payment.student, school)
                        monthly_transport = get_transport_fee(payment.student)

                        exam_fee_items = []
                        exam_total = Decimal('0.00')
                        if selected_exam_keys:
                            exam_pk_list = [int(k.replace('exam_', '')) for k in selected_exam_keys if k.replace('exam_', '').isdigit()]
                            for ef in ExamFee.objects.filter(pk__in=exam_pk_list, school=school):
                                exam_fee_items.append({'name': ef.exam_name, 'amount': str(ef.amount)})
                                exam_total += ef.amount

                        payment.payment_months = selected_months
                        payment.exam_fee_items = exam_fee_items
                        payment.transport_amount = monthly_transport * transport_count
                        payment.amount_paid = (monthly_base * base_count) + (monthly_transport * transport_count) + exam_total
                        payment.gross_amount = payment.amount_paid
                        payment.save()
                        messages.success(request, 'Payment saved.')
                        return redirect('payment_dashboard')
    else:
        form = FeePaymentForm(student_queryset=student_qs, initial_student_id=request.GET.get('student_id'))

    params = request.GET.copy()
    params.pop('page', None)

    selected_student_id = request.GET.get('student_id') or (request.POST.get('student') if request.method == 'POST' else None)
    payment_history = []
    selected_student = None
    advance_available = Decimal('0.00')
    if selected_student_id:
        try:
            selected_student = student_qs.get(pk=selected_student_id)
            payment_history = FeePayment.objects.filter(student=selected_student).select_related('student', 'student__school_class').order_by('-payment_date', '-id')
            advance_available = get_available_advance(selected_student, selected_student.academic_session)
        except Student.DoesNotExist:
            pass

    initial_month_options = get_unpaid_fee_options(selected_student, school) if selected_student else []

    session_class_ids = FeeStructure.objects.filter(
        fee_category__school=school,
        academic_session=current_session,
        frequency=FeeStructure.FREQUENCY_MONTHLY,
    ).values_list('school_class_id', flat=True)
    session_classes = SchoolClass.objects.filter(school=school, pk__in=session_class_ids).order_by('name')

    return render(request, 'school/fees/payment_form.html', {
        'form': form,
        'classes': session_classes,
        'current_session': current_session,
        'filtered_students_page': filtered_page,
        'total_filtered_students': filtered_qs.count(),
        'has_active_filters': has_active_filters,
        'filter_querystring': params.urlencode(),
        'payment_history': payment_history,
        'selected_student': selected_student,
        'initial_month_options': initial_month_options,
        'month_label_map': dict(MONTH_CHOICES),
        'advance_available': advance_available,
    })


@school_only
def payment_detail(request, pk):
    payment = get_object_or_404(FeePayment, pk=pk, school=request.user.school)
    return render(request, 'school/fees/payment_detail.html', {'payment': payment})


@school_only
def payment_edit(request, pk):
    school = request.user.school
    payment = get_object_or_404(FeePayment, pk=pk, school=school)
    student = payment.student
    profile = SchoolProfile.get_for_school(school)
    errors = []

    if request.method == 'POST':
        payment_date_str = request.POST.get('payment_date', '').strip()
        selected_items = request.POST.getlist('payment_months')

        try:
            payment_date = date.fromisoformat(payment_date_str)
        except (ValueError, TypeError):
            errors.append('Invalid payment date.')
            payment_date = payment.payment_date

        if not selected_items:
            errors.append('Select at least one month or fee.')

        if not errors:
            session_months = get_session_months(profile.session_start_month, profile.session_end_month)
            valid_tokens = set(session_months) | {f'{v}_transport' for v in session_months}

            other_paid_tokens = set()
            for month_list in FeePayment.objects.filter(
                student=student, academic_session=payment.academic_session
            ).exclude(pk=pk).values_list('payment_months', flat=True):
                if isinstance(month_list, list):
                    for t in month_list:
                        if str(t) in valid_tokens:
                            other_paid_tokens.add(str(t))

            selected_months = [m for m in selected_items if not str(m).startswith('exam_')]
            selected_exam_keys = [m for m in selected_items if str(m).startswith('exam_')]

            if [m for m in selected_months if m in other_paid_tokens]:
                errors.append('Some selected months are already paid in another payment.')
            else:
                with transaction.atomic():
                    transport_count = sum(1 for m in selected_months if str(m).endswith('_transport'))
                    base_count = len(selected_months) - transport_count
                    monthly_base = get_monthly_tuition_fee(student, school)
                    monthly_transport = get_transport_fee(student)

                    exam_fee_items = []
                    exam_total = Decimal('0.00')
                    if selected_exam_keys:
                        exam_pk_list = [int(k.replace('exam_', '')) for k in selected_exam_keys if k.replace('exam_', '').isdigit()]
                        for ef in ExamFee.objects.filter(pk__in=exam_pk_list, school=school):
                            exam_fee_items.append({'name': ef.exam_name, 'amount': str(ef.amount)})
                            exam_total += ef.amount

                    payment.payment_date = payment_date
                    payment.payment_months = selected_months
                    payment.exam_fee_items = exam_fee_items
                    payment.transport_amount = monthly_transport * transport_count
                    payment.amount_paid = (monthly_base * base_count) + (monthly_transport * transport_count) + exam_total
                    payment.gross_amount = payment.amount_paid
                    payment.save()
                    messages.success(request, 'Payment updated successfully.')
                    return redirect('payment_dashboard')

    month_options = get_editable_fee_options(payment, school)
    student_payments = FeePayment.objects.filter(student=student, school=school).order_by('-payment_date', '-id')
    return render(request, 'school/fees/payment_edit.html', {
        'payment': payment,
        'student': student,
        'month_options': month_options,
        'errors': errors,
        'student_payments': student_payments,
    })


@school_only
def student_fee_structure_ajax(request):
    school = request.user.school
    student_id = request.GET.get('student_id')
    if not student_id:
        return JsonResponse({'items': []})
    student = get_object_or_404(Student, pk=student_id, school=school, status=Student.STATUS_ACTIVE)
    student_session = student.academic_session
    structures = FeeStructure.objects.filter(
        school_class=student.school_class,
        fee_category__school=school,
        fee_category__is_active=True,
        academic_session=student_session,
    ).select_related('fee_category')
    items = [
        {'category_name': r.fee_category.name, 'amount': str(r.amount), 'frequency': r.frequency, 'is_transport': False}
        for r in structures
    ]
    if student.transport_opted and student.transport_amount:
        items.append({'category_name': 'Transport Fee', 'amount': str(student.transport_amount), 'frequency': 'monthly', 'is_transport': True})
    exam_amounts = {
        f'exam_{ef.pk}': float(ef.amount)
        for ef in ExamFee.objects.filter(school=school, school_class=student.school_class, academic_session=student_session)
    }
    advance_available = get_available_advance(student, student_session)
    return JsonResponse({
        'items': items,
        'due_month_options': get_unpaid_fee_options(student, school),
        'student_class': str(student.school_class),
        'session': student_session,
        'has_transport': bool(student.transport_opted and student.transport_amount),
        'advance_available': float(advance_available),
        'exam_amounts': exam_amounts,
    })


@school_only
def lump_sum_preview_ajax(request):
    school = request.user.school
    student_id = request.GET.get('student_id')
    amount_str = request.GET.get('amount', '0').strip()

    if not student_id:
        return JsonResponse({'error': 'No student selected'}, status=400)

    try:
        cash_amount = Decimal(amount_str)
        if cash_amount <= 0:
            raise ValueError
    except (InvalidOperation, ValueError):
        return JsonResponse({'error': 'Invalid amount'}, status=400)

    student = get_object_or_404(Student, pk=student_id, school=school, status=Student.STATUS_ACTIVE)
    profile = SchoolProfile.get_for_school(school)
    session = student.academic_session

    advance_available = get_available_advance(student, session)
    result = distribute_lump_sum(student, school, cash_amount, advance_available, profile)

    monthly_fee = get_monthly_tuition_fee(student, school, session)
    transport_fee = get_transport_fee(student)
    month_label_map = dict(MONTH_CHOICES)
    paid_tokens = set(result['paid_month_tokens'])
    session_months = get_session_months(profile.session_start_month, profile.session_end_month)

    months_breakdown = []
    for month in session_months:
        if month not in paid_tokens:
            continue
        transport_token = f'{month}_transport'
        has_transport_paid = transport_token in paid_tokens
        months_breakdown.append({
            'month': month_label_map.get(month, month.title()),
            'tuition': float(monthly_fee),
            'transport': float(transport_fee) if has_transport_paid else None,
            'row_total': float(monthly_fee + (transport_fee if has_transport_paid else Decimal('0'))),
        })

    return JsonResponse({
        'months_breakdown': months_breakdown,
        'exam_fee_items': result['exam_fee_items'],
        'gross_amount': float(result['gross_amount']),
        'advance_available': float(advance_available),
        'advance_balance': float(result['advance_balance']),
        'cash_amount': float(cash_amount),
        'nothing_to_pay': not months_breakdown and not result['exam_fee_items'],
    })
